"""Find circular references in the latest cutter_v3 output workbook.

A circular reference is a formula that depends on its own cell — directly
(`A1 =A1+1`) or indirectly via another formula on the same sheet. We scan
every formula, build a dependency graph per sheet, then detect cycles.
"""
from __future__ import annotations

import re
import sys
import tempfile
from collections import defaultdict, deque
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string

# Find the freshest cutter_v3 output file in %TEMP%
TEMP_DIR = Path(tempfile.gettempdir())
candidates = sorted(TEMP_DIR.glob("cutter_v3_*.xlsx"),
                    key=lambda p: p.stat().st_mtime, reverse=True)
if not candidates:
    print("No cutter_v3_*.xlsx in %TEMP% yet. Generate one from the UI first.")
    sys.exit(1)

TARGET = candidates[0]
print(f"Inspecting: {TARGET}")
print(f"Size: {TARGET.stat().st_size:,} bytes\n")

wb = load_workbook(TARGET, data_only=False)

# Regex to pull cell references out of a formula. Captures both
#   - same-sheet refs:  A1, $B$10, $C$3
#   - cross-sheet refs:  'Raw data'!$D$2  (we'll skip those for cycle detection)
SAME_SHEET_REF_RE = re.compile(r"(?<!!)\$?([A-Z]{1,3})\$?(\d+)\b")
RANGE_RE = re.compile(r"(?<!!)\$?([A-Z]{1,3})\$?(\d+):\$?([A-Z]{1,3})\$?(\d+)")
CROSS_SHEET_REF_RE = re.compile(r"'?[^'!]+'?!")


def _normalise_ref(s: str) -> str:
    return s.replace("$", "").upper()


def _col_to_idx(letters: str) -> int:
    idx = 0
    for c in letters.upper():
        idx = idx * 26 + (ord(c) - 64)
    return idx


def _idx_to_col(idx: int) -> str:
    out = ""
    while idx > 0:
        idx, rem = divmod(idx - 1, 26)
        out = chr(65 + rem) + out
    return out


def _expand_range(start_col: str, start_row: int, end_col: str, end_row: int) -> list[str]:
    """Expand A1:B3 into [A1, A2, A3, B1, B2, B3]."""
    out = []
    c1 = _col_to_idx(start_col)
    c2 = _col_to_idx(end_col)
    for c in range(c1, c2 + 1):
        for r in range(start_row, end_row + 1):
            out.append(f"{_idx_to_col(c)}{r}")
    return out


def _extract_same_sheet_deps(formula: str, this_sheet: str) -> set[str]:
    """Return cells on this_sheet that this formula depends on."""
    if not formula.startswith("="):
        return set()
    deps: set[str] = set()
    # Walk through the formula, splitting on cross-sheet markers
    parts = re.split(r"('[^']+'!|[A-Za-z_]+\w*!)", formula[1:])
    # parts alternates: text, marker, text, marker, ... — text at even indices
    # is same-sheet content; marker says "next chunk is cross-sheet ref"
    for i, chunk in enumerate(parts):
        if i % 2 == 1:
            # This was a cross-sheet marker. The chunk AFTER is on another sheet.
            # We don't care about cross-sheet refs for cycle detection.
            continue
        # First, gobble ranges (A1:B3)
        for m in RANGE_RE.finditer(chunk):
            c1, r1, c2, r2 = m.group(1), int(m.group(2)), m.group(3), int(m.group(4))
            try:
                for cell in _expand_range(c1, r1, c2, r2):
                    deps.add(_normalise_ref(cell))
            except Exception:
                pass
        # Then, eat single-cell refs that weren't inside ranges
        chunk_no_ranges = RANGE_RE.sub("", chunk)
        for m in SAME_SHEET_REF_RE.finditer(chunk_no_ranges):
            deps.add(f"{m.group(1)}{m.group(2)}")
    return deps


# Build a dep graph per sheet
total_issues = 0
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    graph: dict[str, set[str]] = defaultdict(set)
    cell_formula: dict[str, str] = {}
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v.startswith("="):
                coord = cell.coordinate  # e.g. "B19"
                deps = _extract_same_sheet_deps(v, sheet_name)
                cell_formula[coord] = v
                graph[coord] = deps

    # Find cells whose dep transitively reaches themselves
    def reaches_self(start: str) -> tuple[bool, list[str]]:
        visited: set[str] = set()
        # BFS, tracking path back
        prev: dict[str, str] = {}
        queue: deque = deque([start])
        while queue:
            node = queue.popleft()
            for nxt in graph.get(node, ()):
                if nxt == start:
                    # Reconstruct path
                    path = [start, node]
                    while node in prev:
                        node = prev[node]
                        path.append(node)
                    path.reverse()
                    path.append(start)
                    return True, path
                if nxt in visited:
                    continue
                visited.add(nxt)
                prev[nxt] = node
                queue.append(nxt)
        return False, []

    sheet_issues = 0
    for cell in graph:
        is_circ, path = reaches_self(cell)
        if is_circ:
            sheet_issues += 1
            if sheet_issues <= 4:  # cap output per sheet
                print(f"  [{sheet_name}!{cell}] CIRCULAR via {' -> '.join(path)}")
                print(f"      formula: {cell_formula[cell][:200]}")
    if sheet_issues:
        print(f"  -> {sheet_issues} circular refs on sheet {sheet_name!r}\n")
        total_issues += sheet_issues

print(f"\nTotal circular refs across workbook: {total_issues}")
