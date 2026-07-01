"""Deep dive: dump every formula on every sheet so we can eyeball what Excel
might be flagging. Also list any defined names."""
from __future__ import annotations

import re
import sys
import tempfile
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

TEMP_DIR = Path(tempfile.gettempdir())
candidates = sorted(TEMP_DIR.glob("cutter_v3_*.xlsx"),
                    key=lambda p: p.stat().st_mtime, reverse=True)
TARGET = candidates[0]
print(f"Inspecting: {TARGET}\n")

wb = load_workbook(TARGET, data_only=False)

# 1. Sheet inventory
print("=== Sheet inventory ===")
for sn in wb.sheetnames:
    ws = wb[sn]
    print(f"  {sn:<25}  {ws.max_row:>5} x {ws.max_column:>4}")

# 2. Defined names (Excel global)
print("\n=== Defined names ===")
try:
    names = list(wb.defined_names)
    if not names:
        print("  (none)")
    for n in names[:50]:
        v = wb.defined_names[n].value
        print(f"  {n}: {v}")
except Exception as exc:
    print(f"  (error: {exc})")

# 3. Per-sheet formula dump — show ALL formulas, grouped by sheet
print("\n=== Every formula in the workbook ===")
total = 0
for sn in wb.sheetnames:
    ws = wb[sn]
    sheet_formulas: list[tuple[str, str]] = []
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v.startswith("="):
                sheet_formulas.append((cell.coordinate, v))
    if not sheet_formulas:
        continue
    print(f"\n--- {sn} ({len(sheet_formulas)} formulas) ---")
    # Print first 8 distinct shapes per sheet
    seen_shapes: set[str] = set()
    shown = 0
    for coord, formula in sheet_formulas:
        # Shape = formula with cell refs blanked out
        shape = re.sub(r"\$?[A-Z]{1,3}\$?\d+", "[X]", formula)
        shape = re.sub(r"\s+", " ", shape)[:160]
        if shape in seen_shapes:
            continue
        seen_shapes.add(shape)
        print(f"  [{coord}] {formula[:200].replace(chr(10), ' ')}")
        shown += 1
        if shown >= 8:
            print(f"  … ({len(sheet_formulas) - shown}+ more, shape-grouped)")
            break
    total += len(sheet_formulas)
print(f"\nTotal formulas across workbook: {total}")

# 4. Look for formulas that reference cells in their own sheet AND the same column
#    e.g. `=B19/B23` on a cell in column B — that's the kind of pattern that
#    can cause circular if the dependency closes a loop.
print("\n=== Suspicious patterns (same-column self-area refs) ===")
for sn in wb.sheetnames:
    ws = wb[sn]
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            v = cell.value
            if not isinstance(v, str) or not v.startswith("="):
                continue
            coord = cell.coordinate
            this_col = ''.join(c for c in coord if c.isalpha())
            this_row = int(''.join(c for c in coord if c.isdigit()))
            # Find all same-sheet refs in the formula
            same_sheet_refs = re.findall(
                r"(?<!!)\$?([A-Z]{1,3})\$?(\d+)\b",
                re.sub(r"'?[^'!]+'?!\$?[A-Z]+\$?\d+(?::\$?[A-Z]+\$?\d+)?", "", v),
            )
            for col, row_str in same_sheet_refs:
                if col == this_col and int(row_str) == this_row:
                    print(f"  [{sn}!{coord}] SELF-REF: {v[:150]}")
