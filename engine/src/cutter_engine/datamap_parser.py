"""Spec-compliant datamap parser.

Reads the single-column-A format described in DATAMAP_SPEC.md and produces a
list of `ParsedBlock`. The parser is a small state machine that knows three
states: BETWEEN_BLOCKS, IN_TYPE_HINT, IN_OPTIONS.

Strict: malformed datamaps produce explicit errors rather than guessing.
"""
from __future__ import annotations

import io
import re
from dataclasses import dataclass
from enum import Enum
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .models import ParsedBlock


# ─────────────────────────────────────────────────────────────────────────────
# Regex patterns for the spec
# ─────────────────────────────────────────────────────────────────────────────

# Header row: `[ColumnID]: Question text`  (Col A, col B and C blank)
# Question text may be empty (e.g. piped-from-elsewhere questions like `[Q14]:`).
HEADER_PATTERN = re.compile(r"^\[(?P<id>[^\]]+)\]\s*:\s*(?P<text>.*)$")

# Sub-column row: Col B holds `[SubColID]`
SUB_COLUMN_PATTERN = re.compile(r"^\[(?P<id>[^\]]+)\]\s*$")

# Piping marker in sub-column label (Col C):  `[pipe: helper_col_name]`
# Recognised case-insensitively; the helper col name is captured for inverse-indexing.
PIPE_PATTERN = re.compile(r"^\[\s*pipe\s*:\s*(?P<helper>[^\]]+?)\s*\]\s*$", re.IGNORECASE)

# Type-hint variants
TYPE_HINT_VALUES = re.compile(r"^Values\s*:\s*(?P<lo>-?\d+)\s*-\s*(?P<hi>-?\d+)\s*$", re.IGNORECASE)
TYPE_HINT_OPEN_NUMERIC = re.compile(r"^Open\s+numeric\s+response\s*$", re.IGNORECASE)
TYPE_HINT_OPEN_TEXT = re.compile(r"^Open\s+text\s+response\s*$", re.IGNORECASE)

# Section / comment lines that are visual only and should be ignored
SECTION_DIVIDER_PATTERN = re.compile(r"^\s*(?:>>|Section\s*:)", re.IGNORECASE)
COMMENT_PATTERN = re.compile(r"^\s*#")


# ─────────────────────────────────────────────────────────────────────────────
# Public API
# ─────────────────────────────────────────────────────────────────────────────


class DatamapParseError(ValueError):
    """Raised when the datamap structure is unrecoverable."""


def parse_datamap(path_or_bytes: str | bytes | Path,
                  sheet_name: str | None = None) -> list[ParsedBlock]:
    """Parse a `.xlsx` datamap into a list of ParsedBlock.

    `sheet_name=None` (default) auto-picks the sheet whose name contains
    "datamap"/"map"/"codebook"; falls back to the first sheet.
    """
    content = _read_content(path_or_bytes)
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        sn = sheet_name or _pick_datamap_sheet(wb.sheetnames)
        if sn not in wb.sheetnames:
            raise DatamapParseError(f"Sheet {sn!r} not found in workbook. "
                                    f"Available: {wb.sheetnames}")
        ws = wb[sn]
        rows = _materialise_rows(ws)
    finally:
        wb.close()
    return _parse_rows(rows)


def parse_datamap_from_rows(rows: list[tuple[Any, Any, Any]]) -> list[ParsedBlock]:
    """Parse from pre-materialised (col_a, col_b, col_c) tuples.

    Useful for unit tests and when the datamap is already in memory.
    """
    return _parse_rows(rows)


# ─────────────────────────────────────────────────────────────────────────────
# Implementation
# ─────────────────────────────────────────────────────────────────────────────


def _read_content(path_or_bytes: str | bytes | Path) -> bytes:
    if isinstance(path_or_bytes, (str, Path)):
        return Path(path_or_bytes).read_bytes()
    return path_or_bytes


def _pick_datamap_sheet(sheet_names: list[str]) -> str:
    for keyword in ("datamap", "data map", "codebook", "map", "questions", "metadata"):
        for s in sheet_names:
            if keyword in s.lower():
                return s
    return sheet_names[0] if sheet_names else ""


def _materialise_rows(ws) -> list[tuple[Any, Any, Any]]:
    """Read only the first three columns of every row."""
    out: list[tuple[Any, Any, Any]] = []
    for row in ws.iter_rows(min_row=1, max_col=3, values_only=True):
        # Pad short rows
        row = tuple(row) + (None,) * max(0, 3 - len(row))
        out.append(row[:3])
    return out


class _State(Enum):
    BETWEEN_BLOCKS = "between"
    EXPECT_TYPE_HINT = "expect_type_hint"
    IN_OPTIONS = "in_options"


def _is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return not value.strip()
    return False


def _to_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _parse_rows(rows: list[tuple[Any, Any, Any]]) -> list[ParsedBlock]:
    """The state machine. Returns one ParsedBlock per question block found."""
    blocks: list[ParsedBlock] = []
    state = _State.BETWEEN_BLOCKS

    # Mutable buffers for the in-progress block
    cur_column_id = ""
    cur_question_text = ""
    cur_type_hint = ""
    cur_options: list[tuple[Any, str]] = []
    cur_sub_cols: list[tuple[str, str]] = []
    cur_pipe_sources: list[tuple[str, str]] = []
    cur_warnings: list[str] = []
    cur_source_row = 0

    def flush() -> None:
        if not cur_column_id:
            return
        blocks.append(ParsedBlock(
            column_id=cur_column_id,
            question_text=cur_question_text,
            type_hint=cur_type_hint,
            scale_options=tuple(cur_options),
            sub_columns=tuple(cur_sub_cols),
            pipe_sources=tuple(cur_pipe_sources),
            source_row_in_datamap=cur_source_row,
            warnings=tuple(cur_warnings),
        ))

    def reset() -> None:
        nonlocal cur_column_id, cur_question_text, cur_type_hint
        nonlocal cur_options, cur_sub_cols, cur_pipe_sources, cur_warnings, cur_source_row
        cur_column_id = ""
        cur_question_text = ""
        cur_type_hint = ""
        cur_options = []
        cur_sub_cols = []
        cur_pipe_sources = []
        cur_warnings = []
        cur_source_row = 0

    for row_idx_zero, (a, b, c) in enumerate(rows):
        row_num = row_idx_zero + 1  # 1-indexed for error messages
        a_str = _to_str(a)
        b_str = _to_str(b)
        c_str = _to_str(c)

        # Skip blank rows (and reset state if we were inside a block)
        if _is_blank(a) and _is_blank(b) and _is_blank(c):
            if state != _State.BETWEEN_BLOCKS:
                flush()
                reset()
                state = _State.BETWEEN_BLOCKS
            continue

        # Skip section dividers and comments (Col A only)
        if a_str and (SECTION_DIVIDER_PATTERN.match(a_str) or COMMENT_PATTERN.match(a_str)):
            continue

        # ─ Header row detection: Col A matches "[id]: text" ─
        if a_str:
            header_match = HEADER_PATTERN.match(a_str)
            if header_match:
                # If we were already in a block, that's a missing-blank-row case — flush.
                if state != _State.BETWEEN_BLOCKS:
                    cur_warnings.append(
                        f"No blank-row separator before block at row {row_num} "
                        f"(previous block was {cur_column_id!r})"
                    )
                    flush()
                    reset()
                cur_column_id = header_match.group("id").strip()
                text_part = header_match.group("text").strip()
                # Empty question text is legal (often piped questions like `[Q14]:`).
                # Fall back to the column id so downstream cuts still have a label.
                cur_question_text = text_part if text_part else cur_column_id
                cur_source_row = row_num
                state = _State.EXPECT_TYPE_HINT
                continue
            # Non-header text in Col A while between blocks → parser error
            if state == _State.BETWEEN_BLOCKS:
                raise DatamapParseError(
                    f"Row {row_num}: unrecognised content in column A "
                    f"(not a header `[id]: text`, not a section divider, not a comment): {a_str!r}"
                )
            # Non-header text in Col A while in a block → could be a stray type-hint
            if state == _State.EXPECT_TYPE_HINT:
                cur_type_hint = a_str
                # Validate type-hint format
                if not (TYPE_HINT_VALUES.match(a_str) or
                        TYPE_HINT_OPEN_NUMERIC.match(a_str) or
                        TYPE_HINT_OPEN_TEXT.match(a_str)):
                    cur_warnings.append(
                        f"Row {row_num}: type-hint {a_str!r} is not a recognised pattern "
                        f"(expected 'Values: N-M', 'Open numeric response', or 'Open text response')"
                    )
                state = _State.IN_OPTIONS
                continue
            # Non-header text in Col A while in options → unexpected, warn but stay
            cur_warnings.append(f"Row {row_num}: unexpected non-blank in Col A while reading options: {a_str!r}")
            continue

        # Col A is blank — we should be in EXPECT_TYPE_HINT or IN_OPTIONS
        # Col B holds either an integer code OR a [SubColID]
        if state == _State.EXPECT_TYPE_HINT:
            # Spec violation — type-hint must be in Col A, not in B/C
            raise DatamapParseError(
                f"Row {row_num}: expected type-hint in column A for block {cur_column_id!r}, "
                f"but column A is blank. (got col B={b!r}, col C={c!r})"
            )

        if state == _State.IN_OPTIONS:
            # Sub-column row? `Col B = [SubColID]`
            if b_str:
                sub_match = SUB_COLUMN_PATTERN.match(b_str)
                if sub_match:
                    sub_id = sub_match.group("id").strip()
                    # Detect piped sub-row: label is `[pipe: hQxx_pipingN]`
                    pipe_match = PIPE_PATTERN.match(c_str) if c_str else None
                    if pipe_match:
                        helper_col = pipe_match.group("helper").strip()
                        cur_pipe_sources.append((sub_id, helper_col))
                    cur_sub_cols.append((sub_id, c_str))
                    continue
                # Otherwise interpret B as an option code (int or string)
                code: Any = b_str
                try:
                    code = int(b_str)
                except ValueError:
                    try:
                        code = float(b_str)
                    except ValueError:
                        pass  # leave as string
                cur_options.append((code, c_str))
                continue
            # Col B blank, Col C set → warn
            if c_str:
                cur_warnings.append(f"Row {row_num}: label in Col C with no code in Col B — ignored: {c_str!r}")
            continue

    # End of file — flush trailing block if any
    if state != _State.BETWEEN_BLOCKS:
        flush()
        reset()

    return blocks
