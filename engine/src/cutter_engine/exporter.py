"""Excel exporter — produces the target Bain cuts workbook.

The visual design system is matched cell-for-cell to the reference workbook
(``260223_CE Growth Agenda 2026_cutter_vshare1.xlsx``):

  * Font              : Calibri 11 throughout (bold / italic variants)
  * Section banner    : #C00000 fill, white, **bold + italic**
  * Header / labels   : light-gray fill (#D9D9D9), bold black text, thin borders
  * Code / annotations: gray italic text (#808080) in column A, no fill/border
  * Percentages       : number format ``0.00%``
  * Per-block check    : ``=<total>=SUM(<cells>)`` in gray under each Total row

Column convention on every cut block (mirrors the reference):

    A            B                 C            D ...
    <code>       <row label>       # of resp.   % of resp.
    (gray)       (bold, gray fill) (count)      (=C/total, 0.00%)

Sheet order:
  1.  Output>>            divider
  2.  <Theme 1..N>        theme sheets (Global Filters block + per-question cuts)
  3.  Cross-cuts          ONE sheet — its own Global Filters block, every queued
                          cross-cut stacked with a ``Row x Col`` header + italic
                          full-text subtitle and a live ``#`` / ``%`` matrix
  4.  Mapping>>           divider
  5.  Datamap             verbatim copy of the datamap
  6.  Validation          filter-dropdown lookup tables
  7.  Data>>              divider
  8.  Raw data            verbatim raw data + selection helper columns

All cut formulas reference 'Raw data'; row 1 is the header, data starts at row 2.
"""
from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .models import (
    CrossCutResult,
    FilterSlot,
    QuestionSpec,
    QuestionType,
    Segment,
    SurveySchema,
    ThemeGroup,
)


# ─────────────────────────────────────────────────────────────────────────────
# Styling palette — resolved from the reference workbook
# ─────────────────────────────────────────────────────────────────────────────

_FONT = "Calibri"

_WHITE = "FFFFFF"
_BLACK = "000000"
_RED = "C00000"          # Bain banner red (reference: FFC00000)
_GRAY_FILL = "D9D9D9"    # header / label / total fill (reference lt2 ≈ -0.1)
_GRAY_CODE_FILL = "BFBFBF"  # filter VLOOKUP-code cell (reference lt2 ≈ -0.25)
_GRAY_TEXT = "808080"    # gray annotations in column A / check rows
_GREEN_TEXT = "006100"   # Excel "Good" — max highlight
_GREEN_FILL = "C6EFCE"
_RED_TEXT = "9C0006"     # Excel "Bad" — min highlight
_RED_FILL = "FFC7CE"
_AMBER_FILL = "FFC000"   # WLO "Winners" segmentation banner (future use)

# Fonts ----------------------------------------------------------------------
_F_BANNER = Font(name=_FONT, size=11, bold=True, italic=True, color=_WHITE)
_F_FILTER_NAME = Font(name=_FONT, size=11, color=_WHITE)
_F_HEAD = Font(name=_FONT, size=11, bold=True, color=_BLACK)
_F_LABEL = Font(name=_FONT, size=11, bold=True, color=_BLACK)
_F_BODY = Font(name=_FONT, size=11, color=_BLACK)
_F_ANNOT = Font(name=_FONT, size=11, italic=True, color=_GRAY_TEXT)   # sub-col refs
_F_CODE = Font(name=_FONT, size=11, color=_GRAY_TEXT)                 # numeric codes
_F_CHECK = Font(name=_FONT, size=11, color=_GRAY_TEXT)
_F_SUBTITLE = Font(name=_FONT, size=9, italic=True, color=_GRAY_TEXT) # cross-cut detail
_F_DIV_TITLE = Font(name=_FONT, size=11, bold=True, italic=True, color=_WHITE)
_F_DATAMAP_SECTION = Font(name=_FONT, size=11, bold=True, color=_BLACK)

# Fills ----------------------------------------------------------------------
_FILL_BANNER = PatternFill("solid", fgColor=_RED)
_FILL_GRAY = PatternFill("solid", fgColor=_GRAY_FILL)
_FILL_CODE = PatternFill("solid", fgColor=_GRAY_CODE_FILL)
_FILL_AMBER = PatternFill("solid", fgColor=_AMBER_FILL)

# Borders --------------------------------------------------------------------
_SIDE = Side(style="thin", color=_GRAY_CODE_FILL)
_BORDER = Border(left=_SIDE, right=_SIDE, top=_SIDE, bottom=_SIDE)
# Table framing: thick black outer edge, light-gray thin inner gridlines.
_SIDE_OUTER = Side(style="medium", color=_BLACK)
_SIDE_INNER = Side(style="thin", color=_GRAY_CODE_FILL)


def _table_border(ws, top: int, left: int, bottom: int, right: int) -> None:
    """Frame a rectangular table: medium black outer border, thin gray inner
    gridlines. Applied consistently to the filter panel and every cut block."""
    if bottom < top or right < left:
        return
    for r in range(top, bottom + 1):
        for c in range(left, right + 1):
            ws.cell(row=r, column=c).border = Border(
                top=_SIDE_OUTER if r == top else _SIDE_INNER,
                bottom=_SIDE_OUTER if r == bottom else _SIDE_INNER,
                left=_SIDE_OUTER if c == left else _SIDE_INNER,
                right=_SIDE_OUTER if c == right else _SIDE_INNER,
            )


def _frame_border(ws, top: int, left: int, bottom: int, right: int) -> None:
    """Draw ONLY the outer edge (medium black) around a range — no inner lines,
    so the fills inside read as one contiguous block (no faint 'white' gridlines)."""
    if bottom < top or right < left:
        return
    for r in range(top, bottom + 1):
        for c in range(left, right + 1):
            sides = {}
            if r == top:
                sides["top"] = _SIDE_OUTER
            if r == bottom:
                sides["bottom"] = _SIDE_OUTER
            if c == left:
                sides["left"] = _SIDE_OUTER
            if c == right:
                sides["right"] = _SIDE_OUTER
            if sides:
                ws.cell(row=r, column=c).border = Border(**sides)

# Alignment ------------------------------------------------------------------
_ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)
_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

_PCT_FMT = "0.00%"

# Column layout constants (reference convention) -----------------------------
_CODE_COL = 1     # A — code / annotation (gray)
_LABEL_COL = 2    # B — row label (bold, gray fill)
_METRIC_COL = 3   # C — first metric column


def _style(cell, *, font=None, fill=None, align=None, border=None, numfmt=None):
    """Apply a bundle of style attributes to a cell and return it."""
    if font is not None:
        cell.font = font
    if fill is not None:
        cell.fill = fill
    if align is not None:
        cell.alignment = align
    if border is not None:
        cell.border = border
    if numfmt is not None:
        cell.number_format = numfmt
    return cell


# ─────────────────────────────────────────────────────────────────────────────
# Public API
# ─────────────────────────────────────────────────────────────────────────────


@dataclass(frozen=True, slots=True)
class ExportInputs:
    schema: SurveySchema
    raw_df: pd.DataFrame
    datamap_rows: list[tuple]              # Original datamap rows (col A, B, C) to preserve
    themes: list[ThemeGroup]
    filters: list[FilterSlot]
    cross_cuts: list[CrossCutResult] = ()  # already-computed cross cuts (one stacked sheet)
    segments: list[Segment] = ()           # custom segment filters (helper columns + dropdowns)


_CROSS_CUTS_SHEET = "Cross-cuts"


def export(inputs: ExportInputs, output_path: str | Path) -> Path:
    """Build the workbook and write to `output_path`. Returns the path."""
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    # ── Order of construction matters because some sheets reference others ──
    raw_sheet_name = "Raw data"
    n_raw = len(inputs.raw_df)
    raw_col_to_letter = _build_raw_data_sheet(wb, raw_sheet_name, inputs.raw_df)
    _append_question_sum_helpers(wb, raw_sheet_name, inputs.schema,
                                 raw_col_to_letter, n_raw_rows=n_raw)

    # Segment helper columns (one per segment: labels each respondent with its
    # group name via a live priority IF/AND/OR formula) + an overlap-count column.
    segment_helpers = _append_segment_helpers(
        wb, raw_sheet_name, list(inputs.segments), inputs.schema,
        raw_col_to_letter, n_raw_rows=n_raw)

    _write_divider_sheet(wb, "Output>>")

    # Validation sheet first (theme + cross-cut sheets reference its option lists).
    # Also writes each segment's group-name list for its dropdown.
    filter_validation_ranges = _write_validation_sheet(
        wb, inputs.schema, segment_helpers, n_raw_rows=n_raw)

    # Segments are Global Filters tagged to every cut & cross-cut: append them
    # after the user's question filters so they appear in every filter block.
    all_filters = list(inputs.filters) + [h["filter_slot"] for h in segment_helpers]

    # Theme sheets
    for theme in inputs.themes:
        _write_theme_sheet(wb, theme, inputs.schema, all_filters, raw_col_to_letter,
                           raw_sheet_name, n_raw_rows=n_raw,
                           filter_validation_ranges=filter_validation_ranges)

    # ONE consolidated Cross-cuts sheet (own Global Filters block + live formulas)
    if inputs.cross_cuts:
        _write_cross_cuts_sheet(
            wb, list(inputs.cross_cuts), inputs.schema, all_filters,
            raw_col_to_letter, raw_sheet_name, n_raw_rows=n_raw,
            filter_validation_ranges=filter_validation_ranges,
        )

    _write_divider_sheet(wb, "Mapping>>")
    _write_datamap_sheet(wb, inputs.datamap_rows)

    _write_divider_sheet(wb, "Data>>")

    _reorder_tabs(wb, inputs.themes, bool(inputs.cross_cuts), raw_sheet_name)

    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.calcMode = "auto"
    except Exception:
        pass

    wb.save(out)
    return out


# ─────────────────────────────────────────────────────────────────────────────
# Sheet builders
# ─────────────────────────────────────────────────────────────────────────────


def _safe_sheet(name: str, used: set[str]) -> str:
    cleaned = re.sub(r"[\\/?*\[\]:]", "_", name)[:31] or "Sheet"
    base = cleaned
    i = 1
    while cleaned in used:
        suffix = f"_{i}"
        cleaned = base[: 31 - len(suffix)] + suffix
        i += 1
    used.add(cleaned)
    return cleaned


def _write_divider_sheet(wb: Workbook, name: str) -> None:
    used = set(wb.sheetnames)
    sn = _safe_sheet(name, used)
    ws = wb.create_sheet(sn)
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:D2")
    cell = ws["A1"]
    cell.value = name.replace(">>", "").strip().upper()
    _style(cell, font=_F_DIV_TITLE, fill=_FILL_BANNER,
           align=Alignment(horizontal="center", vertical="center"))
    ws.column_dimensions["A"].width = 24
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 22


def _build_raw_data_sheet(
    wb: Workbook, sheet_name: str, raw_df: pd.DataFrame
) -> dict[str, str]:
    """Write Raw data verbatim. Returns map from column header → Excel column letter."""
    used = set(wb.sheetnames)
    sn = _safe_sheet(sheet_name, used)
    ws = wb.create_sheet(sn)
    ws.sheet_view.showGridLines = False

    col_to_letter: dict[str, str] = {}
    for c_idx, col_name in enumerate(raw_df.columns, start=1):
        letter = get_column_letter(c_idx)
        col_to_letter[str(col_name)] = letter
        _style(ws.cell(row=1, column=c_idx, value=str(col_name)),
               font=_F_HEAD, fill=_FILL_GRAY, border=_BORDER)
        ws.column_dimensions[letter].width = max(12, min(28, len(str(col_name)) + 2))

    for r_idx, (_, row) in enumerate(raw_df.iterrows(), start=2):
        for c_idx, col_name in enumerate(raw_df.columns, start=1):
            v = row[col_name]
            if pd.isna(v):
                continue
            _style(ws.cell(row=r_idx, column=c_idx, value=v), font=_F_BODY)

    ws.freeze_panes = "A2"
    return col_to_letter


def _append_question_sum_helpers(
    wb: Workbook,
    sheet_name: str,
    schema: SurveySchema,
    col_to_letter: dict[str, str],
    n_raw_rows: int,
) -> dict[str, str]:
    """Append a ``_q_sum_<col_id>`` helper column (SUM of all sub-cols per row) for
    every multi-column question (RANKING / MULTI_SELECT_BINARY). Enables a clean
    ``base = COUNTIFS(helper, ">0", ...)`` = respondents who answered the question.
    Mutates `col_to_letter`; returns column_id -> helper_letter.
    """
    ws = wb[sheet_name]
    next_col = ws.max_column + 1
    helpers: dict[str, str] = {}
    SUPPORTED = (QuestionType.RANKING, QuestionType.MULTI_SELECT_BINARY)
    for q in schema.questions:
        if q.question_type not in SUPPORTED:
            continue
        sub_letters = [col_to_letter.get(c) for c in q.raw_columns
                       if col_to_letter.get(c)]
        if not sub_letters:
            continue
        letter = get_column_letter(next_col)
        header = f"_q_sum_{q.column_id}"
        _style(ws.cell(row=1, column=next_col, value=header), font=_F_HEAD, fill=_FILL_GRAY)
        ws.column_dimensions[letter].width = 14
        for row in range(2, n_raw_rows + 2):
            refs = ",".join(f"{l}{row}" for l in sub_letters)
            ws.cell(row=row, column=next_col, value=f"=IFERROR(SUM({refs}),0)")
        col_to_letter[header] = letter
        helpers[q.column_id] = letter
        next_col += 1
    return helpers


def _append_segment_helpers(
    wb: Workbook,
    sheet_name: str,
    segments: list[Segment],
    schema: SurveySchema,
    col_to_letter: dict[str, str],
    n_raw_rows: int,
) -> list[dict]:
    """For each segment, append two live helper columns to Raw data:

      ``_seg_<name>``      — the group label for each respondent (priority IF over
                             AND-of-OR conditions; falls back to the Others label).
      ``_seg_<name>__ovl`` — how many groups that respondent matched (for the
                             overlap warning; should be ≤ 1).

    Returns one dict per segment describing the columns + a ready FilterSlot.
    """
    if not segments:
        return []
    ws = wb[sheet_name]
    out: list[dict] = []
    used_headers: set[str] = set(col_to_letter)

    for seg in segments:
        # Pre-resolve each group's condition expressions (row-independent parts).
        # Each condition -> (colLetter, [(op, value), ...])  OR within predicates.
        valid_groups: list[tuple[str, list[tuple[str, list]]]] = []
        for g in seg.groups:
            conds: list[tuple[str, list]] = []
            for cond in g.conditions:
                # `column` is a raw-data column header; fall back to a question's
                # first raw column for backward compatibility.
                letter = col_to_letter.get(cond.column)
                if letter is None:
                    q = schema.by_column_id(cond.column)
                    if q is not None and q.raw_columns:
                        letter = col_to_letter.get(q.raw_columns[0])
                preds = [(p.op or "=", p.value) for p in cond.predicates
                         if p.value not in (None, "")]
                if not letter or not preds:
                    continue
                conds.append((letter, preds))
            if conds:
                valid_groups.append((g.name, conds))
        if not valid_groups:
            continue

        # Unique headers
        base = re.sub(r"[^0-9A-Za-z]+", "_", seg.name).strip("_") or "segment"
        header = f"_seg_{base}"
        i = 1
        while header in used_headers:
            header = f"_seg_{base}_{i}"; i += 1
        used_headers.add(header)
        ovl_header = f"{header}__ovl"

        assign_col = ws.max_column + 1
        ovl_col = assign_col + 1
        assign_letter = get_column_letter(assign_col)
        ovl_letter = get_column_letter(ovl_col)
        _style(ws.cell(row=1, column=assign_col, value=header), font=_F_HEAD, fill=_FILL_GRAY)
        _style(ws.cell(row=1, column=ovl_col, value=ovl_header), font=_F_HEAD, fill=_FILL_GRAY)
        ws.column_dimensions[assign_letter].width = 16
        ws.column_dimensions[ovl_letter].width = 10

        others = seg.others_label if seg.include_others else ""
        for row in range(2, n_raw_rows + 2):
            matches: list[str] = []
            for _name, conds in valid_groups:
                cond_exprs = []
                for letter, preds in conds:
                    cell = f"${letter}{row}"
                    or_parts = [f"{cell}{op}{_formula_literal(v)}" for op, v in preds]
                    cond_exprs.append("OR(" + ",".join(or_parts) + ")" if len(or_parts) > 1 else or_parts[0])
                matches.append("AND(" + ",".join(cond_exprs) + ")" if len(cond_exprs) > 1 else cond_exprs[0])
            # Nested IF (priority order) → group label, else Others.
            expr = _excel_str(others)
            for (name, _conds), m in zip(reversed(valid_groups), reversed(matches)):
                expr = f'IF({m},{_excel_str(name)},{expr})'
            ws.cell(row=row, column=assign_col, value=f"={expr}")
            ovl = "+".join(f"(--({m}))" for m in matches)
            ws.cell(row=row, column=ovl_col, value=f"={ovl}")

        col_to_letter[header] = assign_letter
        col_to_letter[ovl_header] = ovl_letter
        group_names = [name for name, _ in valid_groups]
        if seg.include_others:
            group_names.append(seg.others_label)
        out.append({
            "segment": seg,
            "helper_header": header,
            "helper_letter": assign_letter,
            "overlap_letter": ovl_letter,
            "group_names": group_names,
            "filter_slot": FilterSlot(name=seg.name, column_id=header, default_value="All"),
        })
    return out


def _write_datamap_sheet(wb: Workbook, datamap_rows: list[tuple]) -> None:
    used = set(wb.sheetnames)
    sn = _safe_sheet("Datamap", used)
    ws = wb.create_sheet(sn)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 65
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 50
    for r_idx, row in enumerate(datamap_rows, start=1):
        a, b, c = (row + (None, None, None))[:3]
        if a is not None:
            cell = ws.cell(row=r_idx, column=1, value=a)
            if isinstance(a, str) and a.startswith("["):
                _style(cell, font=_F_DATAMAP_SECTION, fill=_FILL_GRAY)
            else:
                cell.font = _F_BODY
        if b is not None:
            _style(ws.cell(row=r_idx, column=2, value=b), font=_F_BODY)
        if c is not None:
            _style(ws.cell(row=r_idx, column=3, value=c), font=_F_BODY,
                   align=Alignment(wrap_text=True, vertical="top"))


def _write_validation_sheet(
    wb: Workbook,
    schema: SurveySchema,
    segment_helpers: list[dict] | None = None,
    n_raw_rows: int = 0,
) -> dict[str, tuple[str, str | None]]:
    """Lookup tables that feed the Global Filters dropdowns. Writes label (col B,
    shown in the dropdown) + code (col C, used by COUNTIFS via VLOOKUP).

    Also writes each segment's group-name list (the dropdown values match the
    helper column's text directly, so their lookup_range is None) plus a live
    overlap check counting respondents who matched more than one group.
    Returns column_id -> (label_range, lookup_range | None).
    """
    used = set(wb.sheetnames)
    sn = _safe_sheet("Validation", used)
    ws = wb.create_sheet(sn)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 14
    sheet_name = ws.title

    for col, title in ((1, "Filter"), (2, "Label"), (3, "Code")):
        _style(ws.cell(row=1, column=col, value=title), font=_F_HEAD, fill=_FILL_GRAY, border=_BORDER)
    row = 2
    ranges: dict[str, tuple[str, str]] = {}
    for q in schema.questions:
        if not q.analysis_eligible:
            continue
        if q.question_type != QuestionType.SINGLE_SELECT:
            continue
        if not q.option_map:
            continue
        _style(ws.cell(row=row, column=1, value=f"{q.column_id} — {q.question_text[:40]}"),
               font=_F_DATAMAP_SECTION)
        row += 1
        first_val_row = row
        ws.cell(row=row, column=2, value="All").font = _F_BODY
        ws.cell(row=row, column=3, value="").font = _F_BODY
        row += 1
        for code, label in q.option_map.items():
            ws.cell(row=row, column=2, value=str(label)).font = _F_BODY
            ws.cell(row=row, column=3, value=code).font = _F_BODY
            row += 1
        last_val_row = row - 1
        safe_sn = f"'{sheet_name}'" if any(c in sheet_name for c in " '!") else sheet_name
        label_range = f"{safe_sn}!$B${first_val_row}:$B${last_val_row}"
        lookup_range = f"{safe_sn}!$B${first_val_row}:$C${last_val_row}"
        ranges[q.column_id] = (label_range, lookup_range)
        row += 1  # spacer

    # ── Segment group-name lists (dropdown matches the helper text directly) ──
    for h in (segment_helpers or []):
        seg = h["segment"]
        names = list(h["group_names"])
        _style(ws.cell(row=row, column=1, value=f"Segment: {seg.name}"), font=_F_DATAMAP_SECTION)
        row += 1
        first_val_row = row
        ws.cell(row=row, column=2, value="All").font = _F_BODY
        row += 1
        for nm in names:
            ws.cell(row=row, column=2, value=nm).font = _F_BODY
            row += 1
        last_val_row = row - 1
        safe_sn = f"'{sheet_name}'" if any(c in sheet_name for c in " '!") else sheet_name
        label_range = f"{safe_sn}!$B${first_val_row}:$B${last_val_row}"
        ranges[h["helper_header"]] = (label_range, None)  # None → match name text directly
        # Live overlap check (respondents matching >1 group) — should be 0.
        ovl_letter = h.get("overlap_letter")
        if ovl_letter and n_raw_rows:
            ovl_rng = f"'Raw data'!${ovl_letter}$2:${ovl_letter}${n_raw_rows + 1}"
            _style(ws.cell(row=row, column=1, value="Overlap (should be 0)"), font=_F_ANNOT)
            _style(ws.cell(row=row, column=2, value=f"=SUMPRODUCT(--({ovl_rng}>1))"),
                   font=_F_CHECK)
            row += 1
        row += 1  # spacer

    return ranges


# ── Global Filters block (shared by theme sheets & the Cross-cuts sheet) ──────


def _write_global_filters_block(
    ws,
    filters: list[FilterSlot],
    filter_validation_ranges: dict[str, tuple[str, str]] | None,
) -> list[tuple[FilterSlot, str, str | None]]:
    """Write the reference-style Global Filters block (rows 1..16).

      A1        : "Global Filters" banner (red, bold-italic, white)
      B{r}      : filter name  (red fill, white)
      C{r}      : picked value (gray fill, dropdown)   <- users edit this
      D{r}      : =VLOOKUP(C,label+code,2,0) resolved code (darker gray)

    Returns the last filter row (so callers can frame/divide below it).
    """
    banner = ws.cell(row=1, column=1, value="Global Filters")
    _style(banner, font=_F_BANNER, fill=_FILL_BANNER, align=_ALIGN_LEFT)
    ws.merge_cells("A1:D1")
    for c in range(2, 5):
        _style(ws.cell(row=1, column=c), fill=_FILL_BANNER)
    ws.row_dimensions[1].height = 20

    ranges_map = filter_validation_ranges or {}
    filter_cell_refs: list[tuple[FilterSlot, str, str | None]] = []
    for i, f in enumerate(filters[:12], start=0):
        r = 3 + i
        _style(ws.cell(row=r, column=2, value=f.name),
               font=_F_FILTER_NAME, fill=_FILL_BANNER, align=_ALIGN_LEFT)
        _style(ws.cell(row=r, column=3, value=f.default_value),
               font=_F_BODY, fill=_FILL_GRAY)
        slot_ranges = ranges_map.get(f.column_id) or ranges_map.get(f.name)
        label_range = slot_ranges[0] if slot_ranges else None
        lookup_range = slot_ranges[1] if slot_ranges else None
        if lookup_range:
            _style(ws.cell(row=r, column=4,
                           value=f'=IFERROR(VLOOKUP(C{r},{lookup_range},2,0),"")'),
                   font=_F_BODY, fill=_FILL_CODE)
        else:
            _style(ws.cell(row=r, column=4), fill=_FILL_CODE)
        filter_cell_refs.append((f, f"$C${r}", lookup_range))
        if label_range:
            dv = DataValidation(type="list", formula1=f"={label_range}",
                                allow_blank=True, showDropDown=False)
            dv.add(f"C{r}")
            ws.add_data_validation(dv)
    n = len(filters[:12])
    last_row = 2 + n if n else 1
    # Frame the filter panel with a clean outer border only — no inner lines,
    # so the red/grey fills read as one contiguous block.
    if n:
        _frame_border(ws, 3, 2, last_row, 4)
    return last_row


def _write_theme_sheet(
    wb: Workbook,
    theme: ThemeGroup,
    schema: SurveySchema,
    filters: list[FilterSlot],
    raw_col_to_letter: dict[str, str],
    raw_sheet_name: str,
    n_raw_rows: int,
    filter_validation_ranges: dict[str, tuple[str, str]] | None = None,
) -> None:
    used = set(wb.sheetnames)
    sn = _safe_sheet(theme.name, used)
    ws = wb.create_sheet(sn)
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 76
    for c in range(3, 14):
        ws.column_dimensions[get_column_letter(c)].width = 18

    last_filter_row = _write_global_filters_block(ws, filters, filter_validation_ranges)
    refs = _filter_refs_from_block(filters, filter_validation_ranges)

    # One blank separator row under the filter panel, then the cuts. The filter
    # panel is NOT frozen.
    cur_row = last_filter_row + 2
    for col_id in theme.question_column_ids:
        q = schema.by_column_id(col_id)
        if q is None or not q.analysis_eligible:
            continue
        cur_row = _write_question_cut(
            ws, q, cur_row, refs, raw_col_to_letter, raw_sheet_name, n_raw_rows
        )
        cur_row += 1  # one blank line between question blocks


def _filter_refs_from_block(
    filters: list[FilterSlot],
    filter_validation_ranges: dict[str, tuple[str, str]] | None,
) -> list[tuple[FilterSlot, str, str | None]]:
    """Reconstruct (FilterSlot, "$C$r", lookup_range) for the filter block that
    _write_global_filters_block laid out (rows 3.. in column C)."""
    ranges_map = filter_validation_ranges or {}
    out: list[tuple[FilterSlot, str, str | None]] = []
    for i, f in enumerate(filters[:12], start=0):
        r = 3 + i
        slot_ranges = ranges_map.get(f.column_id) or ranges_map.get(f.name)
        lookup_range = slot_ranges[1] if slot_ranges else None
        out.append((f, f"$C${r}", lookup_range))
    return out


def _write_question_cut(
    ws,
    q: QuestionSpec,
    start_row: int,
    filter_refs: list[tuple[FilterSlot, str, str | None]],
    raw_col_to_letter: dict[str, str],
    raw_sheet_name: str,
    n_raw_rows: int,
) -> int:
    """Write one question's cut block starting at `start_row`. Returns next free row."""
    # Grey question header spanning A:D — bold black on light grey, matching the
    # reference workbook's question row (no red on question headers).
    ws.merge_cells(start_row=start_row, end_row=start_row, start_column=1, end_column=4)
    cell = ws.cell(row=start_row, column=1, value=f"{q.column_id}: {q.question_text}")
    _style(cell, font=_F_LABEL, fill=_FILL_GRAY, align=_ALIGN_LEFT)
    for c in range(2, 5):
        _style(ws.cell(row=start_row, column=c), fill=_FILL_GRAY)
    ws.row_dimensions[start_row].height = 18

    body_row = start_row + 1
    qt = q.question_type

    dispatch = {
        QuestionType.SINGLE_SELECT: _write_single_select_block,
        QuestionType.BINARY_TWO_OPTIONS: _write_single_select_block,
        QuestionType.MULTI_SELECT_BINARY: _write_multi_select_block,
        QuestionType.GRID_RATED: _write_grid_rated_block,
        QuestionType.GRID_SINGLE_SELECT: _write_grid_rated_block,
        QuestionType.NUMERIC_ALLOCATION: _write_numeric_alloc_block,
        QuestionType.NPS: _write_nps_block,
        QuestionType.RANKING: _write_ranking_block,
        QuestionType.DIRECT_NUMERIC: _write_direct_numeric_block,
    }
    builder = dispatch.get(qt)
    if builder is None:
        _style(ws.cell(row=body_row, column=1,
                       value=f"(question type {qt.value} not yet implemented)"), font=_F_BODY)
        return body_row + 1
    return builder(ws, q, body_row, filter_refs, raw_col_to_letter, raw_sheet_name, n_raw_rows)


# ─────────────────────────────────────────────────────────────────────────────
# Per-type formula builders  (A=code · B=label · C..=metrics)
# ─────────────────────────────────────────────────────────────────────────────


def _filter_clause(filter_refs, raw_col_to_letter, raw_sheet_name, n_raw_rows) -> str:
    """COUNTIFS filter clause: for each filter, `range, IF(cell="All","<>", code)`
    where the picked LABEL is translated to its CODE via VLOOKUP against Validation."""
    parts: list[str] = []
    for entry in filter_refs:
        if len(entry) == 3:
            f, cell_ref, lookup_range = entry
        else:
            f, cell_ref = entry
            lookup_range = None
        col_letter = raw_col_to_letter.get(f.column_id)
        if not col_letter:
            continue
        full_range = f"'{raw_sheet_name}'!${col_letter}$2:${col_letter}${n_raw_rows + 1}"
        if lookup_range:
            criterion = (
                f'IF({cell_ref}="All","<>",'
                f'IFERROR(VLOOKUP({cell_ref},{lookup_range},2,FALSE),{cell_ref}))'
            )
        else:
            criterion = f'IF({cell_ref}="All","<>",{cell_ref})'
        parts.append(f"{full_range},{criterion}")
    return ",".join(parts)


def _hdr(ws, row, col, text):
    """Bold header cell on gray fill with border."""
    return _style(ws.cell(row=row, column=col, value=text),
                  font=_F_HEAD, fill=_FILL_GRAY, border=_BORDER, align=_ALIGN_CENTER)


def _label(ws, row, text):
    """Row label in column B — bold on gray fill, bordered."""
    return _style(ws.cell(row=row, column=_LABEL_COL, value=str(text)),
                  font=_F_LABEL, fill=_FILL_GRAY, border=_BORDER, align=_ALIGN_LEFT)


def _code(ws, row, text, *, italic=False):
    """Auxiliary code / annotation in column A — gray, no fill/border."""
    return _style(ws.cell(row=row, column=_CODE_COL, value=text),
                  font=_F_ANNOT if italic else _F_CODE)


def _write_single_select_block(ws, q, start_row, filter_refs, raw_col_to_letter,
                               raw_sheet_name, n_raw_rows) -> int:
    col_letter = raw_col_to_letter.get(q.raw_columns[0]) if q.raw_columns else None
    if col_letter is None:
        _style(ws.cell(row=start_row, column=_LABEL_COL,
                       value=f"(column {q.raw_columns[0] if q.raw_columns else '?'!r} not in raw data)"),
               font=_F_BODY)
        return start_row + 1

    full_range = f"'{raw_sheet_name}'!${col_letter}$2:${col_letter}${n_raw_rows + 1}"
    fclause = _filter_clause(filter_refs, raw_col_to_letter, raw_sheet_name, n_raw_rows)

    # Header row: B ref (italic gray), C "# of respondents", D "% of respondents"
    _code(ws, start_row, q.column_id, italic=True)
    _hdr(ws, start_row, _METRIC_COL, "# of respondents")
    _hdr(ws, start_row, _METRIC_COL + 1, "% of respondents")

    r = start_row + 1
    options = list(q.option_map.items()) if q.option_map else []
    total_row = r + len(options) if options else r + 1
    total_ref = f"$C${total_row}"

    for code, label in options:
        _code(ws, r, code)
        _label(ws, r, label)
        criterion = _excel_str(code)
        formula = (f"=IFERROR(COUNTIFS({full_range},{criterion}"
                   + (f",{fclause}" if fclause else "") + "),0)")
        _style(ws.cell(row=r, column=_METRIC_COL, value=formula), font=_F_BODY, border=_BORDER)
        _style(ws.cell(row=r, column=_METRIC_COL + 1, value=f"=IFERROR(C{r}/{total_ref},0)"),
               font=_F_BODY, border=_BORDER, numfmt=_PCT_FMT)
        r += 1

    # Total respondents row
    _label(ws, r, "Total respondents")
    total_formula = (f"=IFERROR(COUNTIFS({full_range},\"<>\""
                     + (f",{fclause}" if fclause else "") + "),0)")
    _style(ws.cell(row=r, column=_METRIC_COL, value=total_formula),
           font=_F_HEAD, fill=_FILL_GRAY, border=_BORDER)
    first_opt, last_opt = start_row + 1, r - 1
    if last_opt >= first_opt:
        _style(ws.cell(row=r, column=_METRIC_COL + 1,
                       value=f"=IFERROR(SUM(D{first_opt}:D{last_opt}),0)"),
               font=_F_HEAD, fill=_FILL_GRAY, border=_BORDER, numfmt=_PCT_FMT)
        # Check row
        _style(ws.cell(row=r + 1, column=_METRIC_COL,
                       value=f"=C{r}=SUM(C{first_opt}:C{last_opt})"), font=_F_CHECK)
    else:
        _style(ws.cell(row=r, column=_METRIC_COL + 1), fill=_FILL_GRAY, border=_BORDER)
    _table_border(ws, start_row, _LABEL_COL, r, _METRIC_COL + 1)
    return r + 2


def _write_multi_select_block(ws, q, start_row, filter_refs, raw_col_to_letter,
                              raw_sheet_name, n_raw_rows) -> int:
    fclause = _filter_clause(filter_refs, raw_col_to_letter, raw_sheet_name, n_raw_rows)
    _code(ws, start_row, q.column_id, italic=True)
    _hdr(ws, start_row, _METRIC_COL, "# of respondents")
    _hdr(ws, start_row, _METRIC_COL + 1, "% of respondents")

    r = start_row + 1
    if not q.raw_columns:
        _style(ws.cell(row=r, column=_LABEL_COL, value="(no sub-cols in raw data)"), font=_F_BODY)
        return r + 1

    helper_letter = raw_col_to_letter.get(f"_q_sum_{q.column_id}")
    if helper_letter:
        base_range = f"'{raw_sheet_name}'!${helper_letter}$2:${helper_letter}${n_raw_rows + 1}"
        base_criterion = '">0"'
    else:
        first_letter = raw_col_to_letter.get(q.raw_columns[0])
        if not first_letter:
            _style(ws.cell(row=r, column=_LABEL_COL, value="(no sub-cols in raw data)"), font=_F_BODY)
            return r + 1
        base_range = f"'{raw_sheet_name}'!${first_letter}$2:${first_letter}${n_raw_rows + 1}"
        base_criterion = '"<>"'

    base_row = r + len(q.raw_columns)
    base_ref = f"$C${base_row}"

    for sub_col in q.raw_columns:
        letter = raw_col_to_letter.get(sub_col)
        if letter is None:
            continue
        rng = f"'{raw_sheet_name}'!${letter}$2:${letter}${n_raw_rows + 1}"
        _code(ws, r, sub_col, italic=True)
        _label(ws, r, q.sub_column_labels.get(sub_col, sub_col))
        formula = (f"=IFERROR(COUNTIFS({rng},1"
                   + (f",{fclause}" if fclause else "") + "),0)")
        _style(ws.cell(row=r, column=_METRIC_COL, value=formula), font=_F_BODY, border=_BORDER)
        _style(ws.cell(row=r, column=_METRIC_COL + 1, value=f"=IFERROR(C{r}/{base_ref},0)"),
               font=_F_BODY, border=_BORDER, numfmt=_PCT_FMT)
        r += 1

    _label(ws, r, "Base (any answered)")
    _style(ws.cell(row=r, column=_METRIC_COL,
                   value=(f"=IFERROR(COUNTIFS({base_range},{base_criterion}"
                          + (f",{fclause}" if fclause else "") + "),0)")),
           font=_F_HEAD, fill=_FILL_GRAY, border=_BORDER)
    _style(ws.cell(row=r, column=_METRIC_COL + 1), fill=_FILL_GRAY, border=_BORDER)
    _table_border(ws, start_row, _LABEL_COL, r, _METRIC_COL + 1)
    return r + 1


def _write_grid_rated_block(ws, q, start_row, filter_refs, raw_col_to_letter,
                            raw_sheet_name, n_raw_rows) -> int:
    fclause = _filter_clause(filter_refs, raw_col_to_letter, raw_sheet_name, n_raw_rows)
    _code(ws, start_row, q.column_id, italic=True)
    _hdr(ws, start_row, _METRIC_COL, "Valid n")
    _hdr(ws, start_row, _METRIC_COL + 1, "Mean")

    r = start_row + 1
    for sub_col in q.raw_columns:
        letter = raw_col_to_letter.get(sub_col)
        if letter is None:
            continue
        rng = f"'{raw_sheet_name}'!${letter}$2:${letter}${n_raw_rows + 1}"
        _code(ws, r, sub_col, italic=True)
        _label(ws, r, q.sub_column_labels.get(sub_col, sub_col))
        _style(ws.cell(row=r, column=_METRIC_COL,
                       value=(f"=IFERROR(COUNTIFS({rng},\">0\""
                              + (f",{fclause}" if fclause else "") + "),0)")),
               font=_F_BODY, border=_BORDER)
        sum_formula = (f"SUMIFS({rng},{rng},\">0\""
                       + (f",{fclause}" if fclause else "") + ")")
        _style(ws.cell(row=r, column=_METRIC_COL + 1, value=f"=IFERROR({sum_formula}/C{r},0)"),
               font=_F_BODY, border=_BORDER, numfmt="0.00")
        r += 1
    _table_border(ws, start_row, _LABEL_COL, r - 1, _METRIC_COL + 1)
    return r


def _write_numeric_alloc_block(ws, q, start_row, filter_refs, raw_col_to_letter,
                               raw_sheet_name, n_raw_rows) -> int:
    fclause = _filter_clause(filter_refs, raw_col_to_letter, raw_sheet_name, n_raw_rows)
    _code(ws, start_row, q.column_id, italic=True)
    _hdr(ws, start_row, _METRIC_COL, "Valid n")
    _hdr(ws, start_row, _METRIC_COL + 1, "Mean %")

    r = start_row + 1
    for sub_col in q.raw_columns:
        letter = raw_col_to_letter.get(sub_col)
        if letter is None:
            continue
        rng = f"'{raw_sheet_name}'!${letter}$2:${letter}${n_raw_rows + 1}"
        _code(ws, r, sub_col, italic=True)
        _label(ws, r, q.sub_column_labels.get(sub_col, sub_col))
        _style(ws.cell(row=r, column=_METRIC_COL,
                       value=(f"=IFERROR(COUNTIFS({rng},\"<>\""
                              + (f",{fclause}" if fclause else "") + "),0)")),
               font=_F_BODY, border=_BORDER)
        _style(ws.cell(row=r, column=_METRIC_COL + 1,
                       value=(f"=IFERROR(AVERAGEIFS({rng},{rng},\">=0\""
                              + (f",{fclause}" if fclause else "") + "),0)")),
               font=_F_BODY, border=_BORDER, numfmt="0.0")
        r += 1
    _table_border(ws, start_row, _LABEL_COL, r - 1, _METRIC_COL + 1)
    return r


def _write_nps_block(ws, q, start_row, filter_refs, raw_col_to_letter,
                     raw_sheet_name, n_raw_rows) -> int:
    col_letter = raw_col_to_letter.get(q.raw_columns[0]) if q.raw_columns else None
    if col_letter is None:
        _style(ws.cell(row=start_row, column=_LABEL_COL, value="(column missing)"), font=_F_BODY)
        return start_row + 1
    rng = f"'{raw_sheet_name}'!${col_letter}$2:${col_letter}${n_raw_rows + 1}"
    fclause = _filter_clause(filter_refs, raw_col_to_letter, raw_sheet_name, n_raw_rows)

    _code(ws, start_row, q.column_id, italic=True)
    _hdr(ws, start_row, _METRIC_COL, "# of respondents")
    _hdr(ws, start_row, _METRIC_COL + 1, "% of respondents")

    pro_r, pas_r, det_r, nps_r = start_row + 1, start_row + 2, start_row + 3, start_row + 4
    base_ref = f"$C${nps_r + 1}"

    def cnt(row, label, crit):
        _label(ws, row, label)
        _style(ws.cell(row=row, column=_METRIC_COL,
                       value=(f"=IFERROR(COUNTIFS({rng},{crit}"
                              + (f",{fclause}" if fclause else "") + "),0)")),
               font=_F_BODY, border=_BORDER)
        _style(ws.cell(row=row, column=_METRIC_COL + 1, value=f"=IFERROR(C{row}/{base_ref},0)"),
               font=_F_BODY, border=_BORDER, numfmt=_PCT_FMT)

    cnt(pro_r, "Promoters (9-10)", "\">=9\"")
    _label(ws, pas_r, "Passives (7-8)")
    _style(ws.cell(row=pas_r, column=_METRIC_COL,
                   value=(f"=IFERROR(COUNTIFS({rng},\">=7\",{rng},\"<=8\""
                          + (f",{fclause}" if fclause else "") + "),0)")),
           font=_F_BODY, border=_BORDER)
    _style(ws.cell(row=pas_r, column=_METRIC_COL + 1, value=f"=IFERROR(C{pas_r}/{base_ref},0)"),
           font=_F_BODY, border=_BORDER, numfmt=_PCT_FMT)
    cnt(det_r, "Detractors (0-6)", "\"<=6\"")

    _label(ws, nps_r, "NPS")
    _style(ws.cell(row=nps_r, column=_METRIC_COL,
                   value=f"=IFERROR(D{pro_r}*100-D{det_r}*100,0)"),
           font=_F_HEAD, fill=_FILL_GRAY, border=_BORDER, numfmt="+0;-0;0")

    _label(ws, nps_r + 1, "Base")
    _style(ws.cell(row=nps_r + 1, column=_METRIC_COL,
                   value=(f"=IFERROR(COUNTIFS({rng},\">=0\""
                          + (f",{fclause}" if fclause else "") + "),0)")),
           font=_F_HEAD, fill=_FILL_GRAY, border=_BORDER)
    _table_border(ws, start_row, _LABEL_COL, nps_r + 1, _METRIC_COL + 1)
    return nps_r + 2


def _write_ranking_block(ws, q, start_row, filter_refs, raw_col_to_letter,
                         raw_sheet_name, n_raw_rows) -> int:
    """Options on rows, rank positions (1..K) on columns starting at C."""
    fclause = _filter_clause(filter_refs, raw_col_to_letter, raw_sheet_name, n_raw_rows)
    if q.scale_range and q.scale_range[0] is not None:
        rank_lo, rank_hi = q.scale_range
    else:
        rank_lo, rank_hi = 1, len(q.raw_columns) or 1
    rank_values = list(range(int(rank_lo), int(rank_hi) + 1))

    _code(ws, start_row, q.column_id, italic=True)
    for j, rv in enumerate(rank_values):
        _hdr(ws, start_row, _METRIC_COL + j, f"Rank {rv}")

    r = start_row + 1
    for sub_col in q.raw_columns:
        letter = raw_col_to_letter.get(sub_col)
        if letter is None:
            continue
        rng = f"'{raw_sheet_name}'!${letter}$2:${letter}${n_raw_rows + 1}"
        _code(ws, r, sub_col, italic=True)
        _label(ws, r, q.sub_column_labels.get(sub_col, sub_col))
        for j, rv in enumerate(rank_values):
            _style(ws.cell(row=r, column=_METRIC_COL + j,
                           value=(f"=IFERROR(COUNTIFS({rng},{rv}"
                                  + (f",{fclause}" if fclause else "") + "),0)")),
                   font=_F_BODY, border=_BORDER)
        r += 1

    helper_letter = raw_col_to_letter.get(f"_q_sum_{q.column_id}")
    _label(ws, r, "Base (any rank)")
    if helper_letter:
        base_range = f"'{raw_sheet_name}'!${helper_letter}$2:${helper_letter}${n_raw_rows + 1}"
        _style(ws.cell(row=r, column=_METRIC_COL,
                       value=(f"=IFERROR(COUNTIFS({base_range},\">0\""
                              + (f",{fclause}" if fclause else "") + "),0)")),
               font=_F_HEAD, fill=_FILL_GRAY, border=_BORDER)
    _table_border(ws, start_row, _LABEL_COL, r, _METRIC_COL + len(rank_values) - 1)
    return r + 1


def _write_direct_numeric_block(ws, q, start_row, filter_refs, raw_col_to_letter,
                                raw_sheet_name, n_raw_rows) -> int:
    col_letter = raw_col_to_letter.get(q.raw_columns[0]) if q.raw_columns else None
    if col_letter is None:
        _style(ws.cell(row=start_row, column=_LABEL_COL, value="(column missing)"), font=_F_BODY)
        return start_row + 1
    rng = f"'{raw_sheet_name}'!${col_letter}$2:${col_letter}${n_raw_rows + 1}"
    fclause = _filter_clause(filter_refs, raw_col_to_letter, raw_sheet_name, n_raw_rows)

    _code(ws, start_row, q.column_id, italic=True)
    _hdr(ws, start_row, _METRIC_COL, "Statistic")
    _hdr(ws, start_row, _METRIC_COL + 1, "Value")
    _label(ws, start_row + 1, "Mean")
    _style(ws.cell(row=start_row + 1, column=_METRIC_COL + 1,
                   value=(f"=IFERROR(AVERAGEIFS({rng},{rng},\"<>\""
                          + (f",{fclause}" if fclause else "") + "),0)")),
           font=_F_BODY, border=_BORDER, numfmt="0.00")
    _table_border(ws, start_row, _LABEL_COL, start_row + 1, _METRIC_COL + 1)
    return start_row + 2


# ─────────────────────────────────────────────────────────────────────────────
# Consolidated Cross-cuts sheet  (own Global Filters block + live formulas)
# ─────────────────────────────────────────────────────────────────────────────


def _write_cross_cuts_sheet(
    wb: Workbook,
    cross_cuts: list[CrossCutResult],
    schema: SurveySchema,
    filters: list[FilterSlot],
    raw_col_to_letter: dict[str, str],
    raw_sheet_name: str,
    n_raw_rows: int,
    filter_validation_ranges: dict[str, tuple[str, str]] | None = None,
) -> None:
    """ONE 'Cross-cuts' sheet. Global Filters block on top; every cross-cut
    stacked below with a bold ``Row x Col`` header, an italic full-text subtitle,
    and a live ``# / % of respondents`` matrix (COUNTIFS driven by the filters).
    """
    used = set(wb.sheetnames)
    sn = _safe_sheet(_CROSS_CUTS_SHEET, used)
    ws = wb.create_sheet(sn)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 60
    for c in range(3, 30):
        ws.column_dimensions[get_column_letter(c)].width = 16

    last_filter_row = _write_global_filters_block(ws, filters, filter_validation_ranges)
    refs = _filter_refs_from_block(filters, filter_validation_ranges)
    fclause = _filter_clause(refs, raw_col_to_letter, raw_sheet_name, n_raw_rows)

    # One blank separator row under the filters; not frozen.
    r = last_filter_row + 2
    for cc in cross_cuts:
        r = _write_one_cross_cut(ws, cc, schema, raw_col_to_letter, raw_sheet_name,
                                 n_raw_rows, fclause, r)
        r += 2  # spacer between cross-cuts


def _codes_for_axis(q: QuestionSpec | None, labels: tuple[str, ...]) -> list | None:
    """Return the raw CODE for each label of a single-select axis, or None if the
    axis can't be driven by a single raw column + option map (→ static fallback)."""
    if q is None or q.question_type not in (QuestionType.SINGLE_SELECT,
                                            QuestionType.BINARY_TWO_OPTIONS):
        return None
    if not q.option_map or not q.raw_columns:
        return None
    label_to_code = {str(lbl): code for code, lbl in q.option_map.items()}
    codes = [label_to_code.get(str(lbl)) for lbl in labels]
    if any(c is None for c in codes):
        return None
    return codes


def _write_one_cross_cut(ws, cc: CrossCutResult, schema, raw_col_to_letter,
                         raw_sheet_name, n_raw_rows, fclause, start_row) -> int:
    n_cols = len(cc.col_labels)
    if n_cols == 0 or not cc.row_labels:
        return start_row

    count_start = _METRIC_COL                      # C
    count_end = count_start + n_cols - 1
    pct_start = count_end + 1
    pct_end = pct_start + n_cols - 1
    last_col = pct_end

    # ── Header: "Row x Col" (bold, gray fill, spans A..last) ──
    ws.merge_cells(start_row=start_row, end_row=start_row, start_column=1, end_column=last_col)
    _style(ws.cell(row=start_row, column=1, value=f"{cc.row_column_id} x {cc.col_column_id}"),
           font=_F_LABEL, fill=_FILL_GRAY, border=_BORDER, align=_ALIGN_LEFT)
    for c in range(2, last_col + 1):
        _style(ws.cell(row=start_row, column=c), fill=_FILL_GRAY, border=_BORDER)

    # ── Italic full-text subtitle (both questions) ──
    sub_row = start_row + 1
    ws.merge_cells(start_row=sub_row, end_row=sub_row, start_column=1, end_column=last_col)
    _style(ws.cell(row=sub_row, column=1,
                   value=f"{cc.row_column_id}: {cc.row_question_text}  ×  "
                         f"{cc.col_column_id}: {cc.col_question_text}"),
           font=_F_SUBTITLE, align=_ALIGN_LEFT)

    # ── Grouped headers: "# of respondents" / "% of respondents" ──
    grp_row = sub_row + 1
    gh = _hdr(ws, grp_row, count_start, "# of respondents")
    if n_cols > 1:
        ws.merge_cells(start_row=grp_row, end_row=grp_row, start_column=count_start, end_column=count_end)
    for c in range(count_start + 1, count_end + 1):
        _style(ws.cell(row=grp_row, column=c), fill=_FILL_GRAY, border=_BORDER)
    _hdr(ws, grp_row, pct_start, "% of respondents")
    if n_cols > 1:
        ws.merge_cells(start_row=grp_row, end_row=grp_row, start_column=pct_start, end_column=pct_end)
    for c in range(pct_start + 1, pct_end + 1):
        _style(ws.cell(row=grp_row, column=c), fill=_FILL_GRAY, border=_BORDER)

    # ── Column-label sub-header ──
    hdr_row = grp_row + 1
    _code(ws, hdr_row, cc.col_column_id, italic=True)
    for j, lbl in enumerate(cc.col_labels):
        _hdr(ws, hdr_row, count_start + j, str(lbl)[:60])
        _hdr(ws, hdr_row, pct_start + j, str(lbl)[:60])

    # ── Decide live vs static ──
    row_q = schema.by_column_id(cc.row_column_id)
    col_q = schema.by_column_id(cc.col_column_id)
    row_codes = _codes_for_axis(row_q, cc.row_labels)
    col_codes = _codes_for_axis(col_q, cc.col_labels)
    row_letter = raw_col_to_letter.get(row_q.raw_columns[0]) if (row_q and row_q.raw_columns) else None
    col_letter = raw_col_to_letter.get(col_q.raw_columns[0]) if (col_q and col_q.raw_columns) else None
    live = bool(row_codes and col_codes and row_letter and col_letter)

    if not live:
        # Non single-select axis — keep the subtitle, flag the matrix as static.
        _code(ws, hdr_row, f"{cc.col_column_id} (static)", italic=True)

    row_rng = (f"'{raw_sheet_name}'!${row_letter}$2:${row_letter}${n_raw_rows + 1}"
               if row_letter else None)
    col_rng = (f"'{raw_sheet_name}'!${col_letter}$2:${col_letter}${n_raw_rows + 1}"
               if col_letter else None)

    body_start = hdr_row + 1
    n_rows = len(cc.row_labels)
    total_row = body_start + n_rows

    for i, rlbl in enumerate(cc.row_labels):
        rr = body_start + i
        _code(ws, rr, (row_codes[i] if live else i + 1))
        _label(ws, rr, rlbl)
        for j in range(n_cols):
            ccell = ws.cell(row=rr, column=count_start + j)
            if live:
                ccell.value = (f"=IFERROR(COUNTIFS({row_rng},{_excel_str(row_codes[i])},"
                               f"{col_rng},{_excel_str(col_codes[j])}"
                               + (f",{fclause}" if fclause else "") + "),0)")
            else:
                ccell.value = cc.counts[i][j] if i < len(cc.counts) and j < len(cc.counts[i]) else 0
            _style(ccell, font=_F_BODY, border=_BORDER)
            # % = cell / column total
            col_total_ref = f"{get_column_letter(count_start + j)}${total_row}"
            _style(ws.cell(row=rr, column=pct_start + j,
                           value=f"=IFERROR({get_column_letter(count_start + j)}{rr}/{col_total_ref},0)"),
                   font=_F_BODY, border=_BORDER, numfmt=_PCT_FMT)

    # ── Total respondents row ──
    _label(ws, total_row, "Total respondents")
    for j in range(n_cols):
        cl = get_column_letter(count_start + j)
        if live:
            tot = (f"=IFERROR(COUNTIFS({col_rng},{_excel_str(col_codes[j])}"
                   + (f",{fclause}" if fclause else "") + "),0)")
        else:
            tot = sum((cc.counts[i][j] for i in range(n_rows)
                       if i < len(cc.counts) and j < len(cc.counts[i])), 0)
        _style(ws.cell(row=total_row, column=count_start + j, value=tot),
               font=_F_HEAD, fill=_FILL_GRAY, border=_BORDER)
        _style(ws.cell(row=total_row, column=pct_start + j,
                       value=f"=IFERROR(SUM({cl}{body_start}:{cl}{total_row - 1})/{cl}{total_row},0)"),
               font=_F_HEAD, fill=_FILL_GRAY, border=_BORDER, numfmt=_PCT_FMT)

    # ── Check row (each column: total == sum of cells) ──
    check_row = total_row + 1
    for j in range(n_cols):
        cl = get_column_letter(count_start + j)
        _style(ws.cell(row=check_row, column=count_start + j,
                       value=f"={cl}{total_row}=SUM({cl}{body_start}:{cl}{total_row - 1})"),
               font=_F_CHECK)
    # Frame the matrix (grouped headers → total row) with the shared table border.
    _table_border(ws, grp_row, _LABEL_COL, total_row, pct_end)
    return check_row + 1


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────


def _excel_str(value: Any) -> str:
    """Quote a value for use inside an Excel formula."""
    if isinstance(value, str):
        return f'"{value.replace(chr(34), chr(34) * 2)}"'
    return str(value)


def _formula_literal(code: Any) -> str:
    """Literal for a code inside a direct `=` comparison. Numeric codes are
    emitted UNQUOTED (so `A2=1` matches numeric raw data); text codes are quoted.
    COUNTIFS criteria coerce "1"→1, but a bare `=` does not, so this matters."""
    s = str(code).strip()
    try:
        f = float(s)
        return str(int(f)) if f.is_integer() else repr(f)
    except (ValueError, TypeError):
        return f'"{s.replace(chr(34), chr(34) * 2)}"'


def _reorder_tabs(
    wb: Workbook,
    themes: list[ThemeGroup],
    has_cross_cuts: bool,
    raw_sheet_name: str,
) -> None:
    """Tab order: Output>>, themes..., Cross-cuts, Mapping>>, Datamap, Validation,
    Data>>, Raw data."""
    desired: list[str] = ["Output>>"]
    desired.extend(t.name[:31] for t in themes)
    if has_cross_cuts:
        desired.append(_CROSS_CUTS_SHEET)
    desired += ["Mapping>>", "Datamap", "Validation", "Data>>", raw_sheet_name]

    new_order: list[str] = []
    for sn in desired:
        if sn in wb.sheetnames and sn not in new_order:
            new_order.append(sn)
    for sn in wb.sheetnames:
        if sn not in new_order:
            new_order.append(sn)
    wb._sheets = [wb[sn] for sn in new_order]
