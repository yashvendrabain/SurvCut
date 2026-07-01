"""IO helpers — load combined or split datamap+raw, return canonical bytes-safe inputs.

OneDrive Files-on-Demand has been known to lock files transiently; this layer
always reads via a BytesIO buffer so the caller never sees a PermissionError
mid-parse if the source is touched by another process.
"""
from __future__ import annotations

import io
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook


@dataclass(frozen=True, slots=True)
class LoadedInputs:
    raw_df: pd.DataFrame
    datamap_rows: list[tuple]                      # first 3 cols of the datamap sheet
    raw_source_note: str
    datamap_source_note: str


_NORM_RE = re.compile(r"[\s_]+")

def _norm(s: Any) -> str:
    return _NORM_RE.sub("", str(s or "").strip().lower())


def _pick_raw_sheet(wb, exclude: str | None = None) -> str | None:
    KEYWORDS = ("raw data", "rawdata", "raw", "responses", "data sheet")
    for sn in wb.sheetnames:
        if sn == exclude:
            continue
        if any(_norm(k) in _norm(sn) for k in KEYWORDS):
            return sn
    # Fallback: widest sheet that's not the datamap
    candidates = [(sn, wb[sn].max_column or 0) for sn in wb.sheetnames if sn != exclude]
    if not candidates:
        return None
    candidates.sort(key=lambda x: -x[1])
    return candidates[0][0]


def _pick_datamap_sheet(wb) -> str | None:
    KEYWORDS = ("datamap", "data map", "codebook", "questions", "metadata")
    for sn in wb.sheetnames:
        if any(_norm(k) in _norm(sn) for k in KEYWORDS):
            return sn
    return None


def load_combined(path_or_bytes: str | bytes | Path) -> LoadedInputs:
    """Load both sheets from a single combined .xlsx (raw + datamap)."""
    content = _read_bytes(path_or_bytes)
    wb = load_workbook(io.BytesIO(content), read_only=False, data_only=True)
    try:
        dm_sheet = _pick_datamap_sheet(wb)
        raw_sheet = _pick_raw_sheet(wb, exclude=dm_sheet)
        if not dm_sheet:
            raise ValueError(f"Could not find a Datamap sheet in {wb.sheetnames}")
        if not raw_sheet:
            raise ValueError(f"Could not find a Raw data sheet in {wb.sheetnames}")
        if dm_sheet == raw_sheet:
            raise ValueError(f"Datamap and Raw data resolved to the same sheet {dm_sheet!r}")
        datamap_rows = _read_first_three_cols(wb[dm_sheet])
    finally:
        wb.close()
    raw_df = pd.read_excel(io.BytesIO(content), sheet_name=raw_sheet, engine="openpyxl")
    return LoadedInputs(
        raw_df=raw_df,
        datamap_rows=datamap_rows,
        raw_source_note=f"sheet:{raw_sheet}",
        datamap_source_note=f"sheet:{dm_sheet}",
    )


def load_separate(
    raw_path_or_bytes: str | bytes | Path,
    datamap_path_or_bytes: str | bytes | Path,
) -> LoadedInputs:
    """Load a raw .xlsx/.csv and a separate datamap .xlsx."""
    raw_content = _read_bytes(raw_path_or_bytes)
    dm_content = _read_bytes(datamap_path_or_bytes)

    raw_df = _read_first_sheet_as_df(raw_content)
    wb = load_workbook(io.BytesIO(dm_content), read_only=False, data_only=True)
    try:
        dm_sheet = _pick_datamap_sheet(wb) or wb.sheetnames[0]
        datamap_rows = _read_first_three_cols(wb[dm_sheet])
    finally:
        wb.close()
    return LoadedInputs(
        raw_df=raw_df,
        datamap_rows=datamap_rows,
        raw_source_note="(separate raw file)",
        datamap_source_note=f"(separate datamap file)!{dm_sheet}",
    )


def _read_bytes(path_or_bytes: str | bytes | Path) -> bytes:
    if isinstance(path_or_bytes, (str, Path)):
        return Path(path_or_bytes).read_bytes()
    return path_or_bytes


def _read_first_three_cols(ws) -> list[tuple]:
    out: list[tuple] = []
    for row in ws.iter_rows(min_row=1, max_col=3, values_only=True):
        row = tuple(row) + (None,) * max(0, 3 - len(row))
        out.append(row[:3])
    return out


def _read_first_sheet_as_df(content: bytes) -> pd.DataFrame:
    bio = io.BytesIO(content)
    # CSV detection by content sniff: try Excel first, fall back to CSV
    try:
        return pd.read_excel(bio, engine="openpyxl")
    except Exception:
        bio.seek(0)
        return pd.read_csv(bio)
