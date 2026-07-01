"""Excel exporter — produces the target Bain cuts workbook.

Sheet order matches the reference file:
  1.  Output>>            divider
  2.  <Theme 1>           theme sheet with Global Filters block + per-question cuts
  3.  <Theme 2>
  4.  ...
  N.  Mapping>>           divider
  N+1. Datamap            verbatim copy of the user's datamap
  N+2. Validation         filter dropdown lookups (auto-generated)
  N+3. Data>>             divider
  N+4. Raw data           verbatim copy of the user's raw data + selection helpers
  N+5..N+M. <CrossCut N>  one sheet per user-picked cross cut

Each theme sheet has:
  Rows 1–14:  Global Filters block (dropdowns that drive all cuts on this sheet)
  Row 16+:    For each question in the theme:
                - question header
                - cuts (COUNTIFS / SUMIFS / AVERAGEIFS with the global filters)

All cut formulas reference 'Raw data' via direct $col$row ranges. The first
row of 'Raw data' is the header; data starts at row 2.
"""
from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .models import (
    CrossCutResult,
    FilterSlot,
    QuestionSpec,
    QuestionType,
    SurveySchema,
    ThemeGroup,
)


# ─────────────────────────────────────────────────────────────────────────────
# Styling palette (Bain red)
# ─────────────────────────────────────────────────────────────────────────────

_F_DIV_TITLE = Font(name="Arial", size=20, bold=True, color="FFFFFF")
_FILL_DIVIDER = PatternFill("solid", fgColor="CC0000")

_F_THEME_TITLE = Font(name="Arial", size=14, bold=True, color="0A0A0A")
_F_SECTION = Font(name="Arial", size=11, bold=True, color="0A0A0A")
_FILL_SECTION = PatternFill("solid", fgColor="F2F2F2")

_F_HEAD = Font(name="Arial", size=10, bold=True, color="0A0A0A")
_FILL_HEAD = PatternFill("solid", fgColor="E0E0E0")

_F_QHEAD = Font(name="Arial", size=11, bold=True, color="FFFFFF")
_FILL_QHEAD = PatternFill("solid", fgColor="0A0A0A")

_F_BODY = Font(name="Arial", size=10, color="0A0A0A")
_F_PCT = Font(name="Arial", size=10, color="0A0A0A", italic=True)

_BORDER_THIN = Border(
    left=Side(style="thin", color="C0C0C0"),
    right=Side(style="thin", color="C0C0C0"),
    top=Side(style="thin", color="C0C0C0"),
    bottom=Side(style="thin", color="C0C0C0"),
)

ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")


# ─────────────────────────────────────────────────────────────────────────────
# Public API
# ─────────────────────────────────────────────────────────────────────────────


@dataclass(frozen=True, slots=True)
class ExportInputs:
    schema: SurveySchema
    raw_df: pd.DataFrame
    datamap_rows: list[tuple]              # Original datamap rows (col A, B, C) to preserve
    themes: list[ThemeGroup]
    filters: list[FilterSlot]
    cross_cuts: list[CrossCutResult] = ()  # already-computed cross cuts to include as own sheets


def export(inputs: ExportInputs, output_path: str | Path) -> Path:
    """Build the workbook and write to `output_path`. Returns the path."""
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    # ── Order of construction matters because some sheets reference others ──
    raw_sheet_name = "Raw data"
    raw_col_to_letter = _build_raw_data_sheet(wb, raw_sheet_name, inputs.raw_df)
    # Append SUM helper columns for RANKING and MULTI_SELECT_BINARY questions.
    # Headers: `_q_sum_<col_id>`. Used by their base formulas.
    _append_question_sum_helpers(wb, raw_sheet_name, inputs.schema,
                                  raw_col_to_letter, n_raw_rows=len(inputs.raw_df))

    # Output>> divider
    _write_divider_sheet(wb, "Output>>")

    # Validation sheet first (theme sheets reference its option lists via
    # data-validation dropdowns; tab order is fixed up at the end by _reorder_tabs)
    filter_validation_ranges = _write_validation_sheet(wb, inputs.schema)

    # Theme sheets
    for theme in inputs.themes:
        _write_theme_sheet(wb, theme, inputs.schema, inputs.filters, raw_col_to_letter,
                           raw_sheet_name, n_raw_rows=len(inputs.raw_df),
                           filter_validation_ranges=filter_validation_ranges)

    # Mapping>> divider + Datamap (Validation already written above)
    _write_divider_sheet(wb, "Mapping>>")
    _write_datamap_sheet(wb, inputs.datamap_rows)

    # Data>> divider — Raw data was already created up top, but Excel
    # tab order is rebuilt explicitly below.
    _write_divider_sheet(wb, "Data>>")

    # Cross cut sheets (user-picked)
    for cc in inputs.cross_cuts:
        _write_cross_cut_sheet(wb, cc)

    # Re-order tabs to match the reference layout
    _reorder_tabs(wb, inputs.themes, inputs.cross_cuts, raw_sheet_name)

    # Force Excel to recalc all formulas on open
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.calcMode = "auto"
    except Exception:
        pass

    wb.save(out)
    return out


# ─────────────────────────────────────────────────────────────────────────────
# Sheet builders
# ─────────────────────────────────────────────────────────────────────────────


def _safe_sheet(name: str, used: set[str]) -> str:
    cleaned = re.sub(r"[\\/?*\[\]:]", "_", name)[:31] or "Sheet"
    base = cleaned
    i = 1
    while cleaned in used:
        suffix = f"_{i}"
        cleaned = base[: 31 - len(suffix)] + suffix
        i += 1
    used.add(cleaned)
    return cleaned


def _write_divider_sheet(wb: Workbook, name: str) -> None:
    used = set(wb.sheetnames)
    sn = _safe_sheet(name, used)
    ws = wb.create_sheet(sn)
    ws.merge_cells("A1:D2")
    cell = ws["A1"]
    cell.value = name.replace(">>", "").strip().upper()
    cell.font = _F_DIV_TITLE
    cell.fill = _FILL_DIVIDER
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions["A"].width = 30
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 30


def _build_raw_data_sheet(
    wb: Workbook, sheet_name: str, raw_df: pd.DataFrame
) -> dict[str, str]:
    """Write Raw data verbatim. Returns map from column header → Excel column letter."""
    used = set(wb.sheetnames)
    sn = _safe_sheet(sheet_name, used)
    ws = wb.create_sheet(sn)

    col_to_letter: dict[str, str] = {}
    for c_idx, col_name in enumerate(raw_df.columns, start=1):
        letter = get_column_letter(c_idx)
        col_to_letter[str(col_name)] = letter
        ws.cell(row=1, column=c_idx, value=str(col_name)).font = _F_HEAD
        ws.cell(row=1, column=c_idx).fill = _FILL_HEAD
        ws.column_dimensions[letter].width = max(12, min(28, len(str(col_name)) + 2))

    for r_idx, (_, row) in enumerate(raw_df.iterrows(), start=2):
        for c_idx, col_name in enumerate(raw_df.columns, start=1):
            v = row[col_name]
            if pd.isna(v):
                continue
            ws.cell(row=r_idx, column=c_idx, value=v)

    ws.freeze_panes = "A2"
    return col_to_letter


def _append_question_sum_helpers(
    wb: Workbook,
    sheet_name: str,
    schema: SurveySchema,
    col_to_letter: dict[str, str],
    n_raw_rows: int,
) -> dict[str, str]:
    """For each multi-column question (RANKING or MULTI_SELECT_BINARY), append a
    derived `_q_sum_<col_id>` column at the right edge of the Raw data sheet
    whose value is SUM(all sub-cols) per row.

    The helper enables a clean, correct "base = respondents who participated in
    this question" formula:
        =COUNTIFS(helper_col, ">0", <filters>)

    For RANKING, sub-cols hold integer ranks (1..K); a respondent who ranked
    any item has SUM > 0, one who skipped has 0.

    For MULTI_SELECT_BINARY, sub-cols hold 0/1; a respondent who picked any
    option has SUM > 0, one who picked none has 0.

    Without the helper, the base falls back to "respondent answered the FIRST
    sub-col" which silently undercounts whenever respondents skip that
    specific sub-col.

    Mutates `col_to_letter` (adds the helper headers) and returns
    column_id -> helper_letter for the questions that got a helper.
    """
    ws = wb[sheet_name]
    next_col = ws.max_column + 1
    helpers: dict[str, str] = {}
    SUPPORTED = (QuestionType.RANKING, QuestionType.MULTI_SELECT_BINARY)
    for q in schema.questions:
        if q.question_type not in SUPPORTED:
            continue
        sub_letters = [col_to_letter.get(c) for c in q.raw_columns
                       if col_to_letter.get(c)]
        if not sub_letters:
            continue
        letter = get_column_letter(next_col)
        header = f"_q_sum_{q.column_id}"
        ws.cell(row=1, column=next_col, value=header).font = _F_HEAD
        ws.cell(row=1, column=next_col).fill = _FILL_HEAD
        ws.column_dimensions[letter].width = 14
        for row in range(2, n_raw_rows + 2):
            refs = ",".join(f"{l}{row}" for l in sub_letters)
            ws.cell(row=row, column=next_col, value=f"=IFERROR(SUM({refs}),0)")
        col_to_letter[header] = letter
        helpers[q.column_id] = letter
        next_col += 1
    return helpers


def _write_datamap_sheet(wb: Workbook, datamap_rows: list[tuple]) -> None:
    used = set(wb.sheetnames)
    sn = _safe_sheet("Datamap", used)
    ws = wb.create_sheet(sn)
    ws.column_dimensions["A"].width = 65
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 50
    for r_idx, row in enumerate(datamap_rows, start=1):
        a, b, c = (row + (None, None, None))[:3]
        if a is not None:
            cell = ws.cell(row=r_idx, column=1, value=a)
            if isinstance(a, str) and a.startswith("["):
                cell.font = _F_SECTION
                cell.fill = _FILL_SECTION
        if b is not None:
            ws.cell(row=r_idx, column=2, value=b)
        if c is not None:
            ws.cell(row=r_idx, column=3, value=c).alignment = Alignment(wrap_text=True, vertical="top")


def _write_validation_sheet(wb: Workbook, schema: SurveySchema) -> dict[str, tuple[str, str]]:
    """Lookup tables that feed the Global Filters dropdowns on each theme sheet.

    Writes BOTH the label (col B, visible to user via dropdown) AND the
    underlying option code (col C, used by COUNTIFS via VLOOKUP).

    Returns a dict: column_id -> (label_range, lookup_range)
      label_range  = "Validation!$B$3:$B$15"   (dropdown values — labels only)
      lookup_range = "Validation!$B$3:$C$15"   (label+code — VLOOKUP source)
    """
    used = set(wb.sheetnames)
    sn = _safe_sheet("Validation", used)
    ws = wb.create_sheet(sn)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 14
    sheet_name = ws.title  # may be deduped by _safe_sheet

    ws.cell(row=1, column=1, value="Filter").font = _F_HEAD
    ws.cell(row=1, column=2, value="Label").font = _F_HEAD
    ws.cell(row=1, column=3, value="Code").font = _F_HEAD
    row = 2
    ranges: dict[str, tuple[str, str]] = {}
    # For every single-select question with options, write a 2-col lookup table
    # (label, code). The dropdown points at the label column only; the COUNTIFS
    # criterion does a VLOOKUP(label) -> code so it can match the numeric codes
    # actually present in raw data.
    for q in schema.questions:
        if not q.analysis_eligible:
            continue
        if q.question_type != QuestionType.SINGLE_SELECT:
            continue
        if not q.option_map:
            continue
        ws.cell(row=row, column=1, value=f"{q.column_id} — {q.question_text[:40]}").font = _F_SECTION
        row += 1
        first_val_row = row
        # "All" row — special sentinel, no code needed (the IF check handles it
        # before VLOOKUP runs).
        ws.cell(row=row, column=2, value="All")
        ws.cell(row=row, column=3, value="")
        row += 1
        for code, label in q.option_map.items():
            ws.cell(row=row, column=2, value=str(label))
            ws.cell(row=row, column=3, value=code)
            row += 1
        last_val_row = row - 1
        # Quote sheet name if it contains spaces or special chars
        safe_sn = f"'{sheet_name}'" if any(c in sheet_name for c in " '!") else sheet_name
        label_range = f"{safe_sn}!$B${first_val_row}:$B${last_val_row}"
        lookup_range = f"{safe_sn}!$B${first_val_row}:$C${last_val_row}"
        ranges[q.column_id] = (label_range, lookup_range)
        row += 1  # spacer between filters

    return ranges


def _write_theme_sheet(
    wb: Workbook,
    theme: ThemeGroup,
    schema: SurveySchema,
    filters: list[FilterSlot],
    raw_col_to_letter: dict[str, str],
    raw_sheet_name: str,
    n_raw_rows: int,
    filter_validation_ranges: dict[str, tuple[str, str]] | None = None,
) -> None:
    used = set(wb.sheetnames)
    sn = _safe_sheet(theme.name, used)
    ws = wb.create_sheet(sn)

    # Column widths
    ws.column_dimensions["A"].width = 50
    for c in range(2, 12):
        ws.column_dimensions[get_column_letter(c)].width = 16

    # ── Master Check (A1/B1) — universal sanity flag per §10.2 of CUTS_FRAMEWORK.md
    # B1 will become =AND(<every block's True/False cell>) once per-block validation
    # rows are wired up. For Phase 1.4 a literal TRUE serves as a placeholder so the
    # cell exists in the expected position.
    ws.cell(row=1, column=1, value="Master Check").font = _F_HEAD
    ws.cell(row=1, column=2, value=True).font = _F_HEAD

    # ── Global Filters block (title at row 2, filter rows start at row 3) ──
    ws.cell(row=2, column=1, value="Global Filters").font = _F_THEME_TITLE
    ws.cell(row=2, column=1).fill = _FILL_HEAD
    ws.row_dimensions[2].height = 22

    # Each filter takes one row starting at row 3.
    # filter_cell_refs items are (FilterSlot, cell_ref, lookup_range_or_None).
    # The 3rd element is the VLOOKUP source range for label -> code translation
    # used inside COUNTIFS criteria.
    filter_cell_refs: list[tuple[FilterSlot, str, str | None]] = []
    ranges_map = filter_validation_ranges or {}
    for i, f in enumerate(filters[:12], start=0):
        r = 3 + i
        ws.cell(row=r, column=2, value=f.name).font = _F_HEAD
        ws.cell(row=r, column=3, value=f.default_value)  # the cell users edit
        slot_ranges = ranges_map.get(f.column_id) or ranges_map.get(f.name)
        label_range = slot_ranges[0] if slot_ranges else None
        lookup_range = slot_ranges[1] if slot_ranges else None
        filter_cell_refs.append((f, f"$C${r}", lookup_range))
        # Attach a data validation dropdown sourced from the label column only.
        if label_range:
            dv = DataValidation(type="list", formula1=f"={label_range}",
                                allow_blank=True, showDropDown=False)
            dv.add(f"C{r}")
            ws.add_data_validation(dv)

    # ── Per-question cuts (start at row 17) ──
    cur_row = 17
    for col_id in theme.question_column_ids:
        q = schema.by_column_id(col_id)
        if q is None or not q.analysis_eligible:
            continue
        cur_row = _write_question_cut(
            ws, q, cur_row, filter_cell_refs, raw_col_to_letter, raw_sheet_name, n_raw_rows
        )
        cur_row += 2  # spacer after each question

    ws.freeze_panes = "A17"


def _write_question_cut(
    ws,
    q: QuestionSpec,
    start_row: int,
    filter_refs: list[tuple[FilterSlot, str]],
    raw_col_to_letter: dict[str, str],
    raw_sheet_name: str,
    n_raw_rows: int,
) -> int:
    """Write one question's cut block starting at `start_row`. Returns next free row."""
    # Header
    ws.merge_cells(start_row=start_row, end_row=start_row, start_column=1, end_column=4)
    cell = ws.cell(row=start_row, column=1, value=f"{q.column_id}: {q.question_text}")
    cell.font = _F_QHEAD
    cell.fill = _FILL_QHEAD
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[start_row].height = 20

    body_row = start_row + 1
    qt = q.question_type

    if qt == QuestionType.SINGLE_SELECT:
        body_row = _write_single_select_block(
            ws, q, body_row, filter_refs, raw_col_to_letter, raw_sheet_name, n_raw_rows
        )
    elif qt == QuestionType.MULTI_SELECT_BINARY:
        body_row = _write_multi_select_block(
            ws, q, body_row, filter_refs, raw_col_to_letter, raw_sheet_name, n_raw_rows
        )
    elif qt == QuestionType.GRID_RATED:
        body_row = _write_grid_rated_block(
            ws, q, body_row, filter_refs, raw_col_to_letter, raw_sheet_name, n_raw_rows
        )
    elif qt == QuestionType.NUMERIC_ALLOCATION:
        body_row = _write_numeric_alloc_block(
            ws, q, body_row, filter_refs, raw_col_to_letter, raw_sheet_name, n_raw_rows
        )
    elif qt == QuestionType.NPS:
        body_row = _write_nps_block(
            ws, q, body_row, filter_refs, raw_col_to_letter, raw_sheet_name, n_raw_rows
        )
    elif qt == QuestionType.RANKING:
        body_row = _write_ranking_block(
            ws, q, body_row, filter_refs, raw_col_to_letter, raw_sheet_name, n_raw_rows
        )
    elif qt == QuestionType.DIRECT_NUMERIC:
        body_row = _write_direct_numeric_block(
            ws, q, body_row, filter_refs, raw_col_to_letter, raw_sheet_name, n_raw_rows
        )
    elif qt == QuestionType.BINARY_TWO_OPTIONS:
        body_row = _write_single_select_block(
            ws, q, body_row, filter_refs, raw_col_to_letter, raw_sheet_name, n_raw_rows
        )
    else:
        ws.cell(row=body_row, column=1,
                value=f"(question type {qt.value} not yet implemented)").font = _F_BODY
        body_row += 1
    return body_row


# ─────────────────────────────────────────────────────────────────────────────
# Per-type formula builders
# ─────────────────────────────────────────────────────────────────────────────


def _filter_clause(filter_refs: list[tuple[FilterSlot, str]],
                   raw_col_to_letter: dict[str, str],
                   raw_sheet_name: str, n_raw_rows: int) -> str:
    """Build the COUNTIFS filter clause for the global filters block.

    Skip filters whose value is "All" — implemented via IF(value="All", ..., specific).
    For simplicity & robustness across Excel versions, we ONLY include filters
    that target an existing raw column AND use a value other than the default "All".
    We always include the filter so the user can change them later, but each
    filter is wrapped with `IF(cell="All", any-match, exact-match)` via a
    sentinel range trick: we pass the cell value directly and let COUNTIFS treat
    "All" literally — combined with a chained IF in the cell, the user just
    edits the cell to filter.

    PRACTICAL CHOICE: We embed each filter as `range, IF(cell="All", "*", cell)`.
    "*" in COUNTIFS matches any non-blank, which acts as a pass-through.
    """
    parts: list[str] = []
    for entry in filter_refs:
        # Backward compatible: accept 2- or 3-tuple
        if len(entry) == 3:
            f, cell_ref, lookup_range = entry
        else:
            f, cell_ref = entry
            lookup_range = None
        col_letter = raw_col_to_letter.get(f.column_id)
        if not col_letter:
            continue
        full_range = f"'{raw_sheet_name}'!${col_letter}$2:${col_letter}${n_raw_rows + 1}"
        # When the dropdown is "All" -> "<>" (any non-blank pass-through).
        # Otherwise translate the picked LABEL to its numeric CODE via VLOOKUP
        # against the Validation sheet, so COUNTIFS matches the numeric data in
        # raw data. Without VLOOKUP, picking a label like "Female" would compare
        # the text "Female" against numeric codes (1, 2, 3...) and find no rows.
        if lookup_range:
            criterion = (
                f'IF({cell_ref}="All","<>",'
                f'IFERROR(VLOOKUP({cell_ref},{lookup_range},2,FALSE),{cell_ref}))'
            )
        else:
            criterion = f'IF({cell_ref}="All","<>",{cell_ref})'
        parts.append(f"{full_range},{criterion}")
    return ",".join(parts)


def _write_single_select_block(ws, q, start_row, filter_refs, raw_col_to_letter,
                                raw_sheet_name, n_raw_rows) -> int:
    col_letter = raw_col_to_letter.get(q.raw_columns[0])
    if col_letter is None:
        ws.cell(row=start_row, column=1, value=f"(column {q.raw_columns[0]!r} not in raw data)")
        return start_row + 1

    full_range = f"'{raw_sheet_name}'!${col_letter}$2:${col_letter}${n_raw_rows + 1}"
    fclause = _filter_clause(filter_refs, raw_col_to_letter, raw_sheet_name, n_raw_rows)

    # Header row
    ws.cell(row=start_row, column=1, value="Option").font = _F_HEAD
    ws.cell(row=start_row, column=1).fill = _FILL_HEAD
    ws.cell(row=start_row, column=2, value="Count").font = _F_HEAD
    ws.cell(row=start_row, column=2).fill = _FILL_HEAD
    ws.cell(row=start_row, column=3, value="%").font = _F_HEAD
    ws.cell(row=start_row, column=3).fill = _FILL_HEAD

    r = start_row + 1
    # Total formula (used in % denominator) — first writes per-option rows, then a total row
    total_row = r + len(q.option_map) if q.option_map else r + 1
    total_ref = f"$B${total_row}"

    # Per-option rows
    options = list(q.option_map.items()) if q.option_map else []
    for code, label in options:
        ws.cell(row=r, column=1, value=str(label)).font = _F_BODY
        # COUNTIFS: data col = code, then all filters
        criterion = _excel_str(code)
        formula = (
            f"=IFERROR(COUNTIFS({full_range},{criterion}"
            + (f",{fclause}" if fclause else "")
            + "),0)"
        )
        ws.cell(row=r, column=2, value=formula).font = _F_BODY
        ws.cell(row=r, column=3, value=f"=IFERROR(B{r}/{total_ref},0)").number_format = "0.0%"
        ws.cell(row=r, column=3).font = _F_PCT
        r += 1

    # Total row
    ws.cell(row=r, column=1, value="Total").font = _F_HEAD
    ws.cell(row=r, column=1).fill = _FILL_HEAD
    # Total = COUNTIFS with "<>" (any non-blank) + filters
    total_formula = (
        f"=IFERROR(COUNTIFS({full_range},\"<>\""
        + (f",{fclause}" if fclause else "")
        + "),0)"
    )
    ws.cell(row=r, column=2, value=total_formula).font = _F_HEAD
    ws.cell(row=r, column=2).fill = _FILL_HEAD
    # Percent total ONLY when there are actually option rows above to sum.
    # If `options` was empty, the per-option loop didn't run, so r == start_row + 1,
    # and a SUM(C{start_row+1}:C{r-1}) is a reversed range that Excel treats as a
    # self-reference (the cell holding the formula is C{r} == C{start_row+1}).
    # That triggers a circular-reference warning on open.
    first_option_row = start_row + 1
    last_option_row = r - 1
    if last_option_row >= first_option_row:
        ws.cell(row=r, column=3,
                value=f"=IFERROR(SUM(C{first_option_row}:C{last_option_row}),0)"
                ).number_format = "0.0%"
        ws.cell(row=r, column=3).fill = _FILL_HEAD
    else:
        # No options were declared in the datamap — leave the % blank, write
        # only the count in B{r}. Excel won't have anything to flag.
        ws.cell(row=r, column=3).fill = _FILL_HEAD
    return r + 1


def _write_multi_select_block(ws, q, start_row, filter_refs, raw_col_to_letter,
                                raw_sheet_name, n_raw_rows) -> int:
    fclause = _filter_clause(filter_refs, raw_col_to_letter, raw_sheet_name, n_raw_rows)
    ws.cell(row=start_row, column=1, value="Option").font = _F_HEAD
    ws.cell(row=start_row, column=1).fill = _FILL_HEAD
    ws.cell(row=start_row, column=2, value="# Selected").font = _F_HEAD
    ws.cell(row=start_row, column=2).fill = _FILL_HEAD
    ws.cell(row=start_row, column=3, value="% of Base").font = _F_HEAD
    ws.cell(row=start_row, column=3).fill = _FILL_HEAD

    r = start_row + 1
    if not q.raw_columns:
        ws.cell(row=r, column=1, value="(no sub-cols in raw data)")
        return r + 1

    # Base = respondents who selected at least ONE option in this question.
    # Uses the `_q_sum_<col_id>` helper column (SUM of all sub-cols per row),
    # so COUNTIFS(helper, ">0") cleanly counts respondents who picked any
    # option. Without the helper, the base would fall back to the first
    # sub-col which silently excluded respondents who didn't pick that one.
    helper_header = f"_q_sum_{q.column_id}"
    helper_letter = raw_col_to_letter.get(helper_header)
    if helper_letter:
        base_range = f"'{raw_sheet_name}'!${helper_letter}$2:${helper_letter}${n_raw_rows + 1}"
        base_criterion = '">0"'
    else:
        # Fallback (shouldn't trigger post-helper-rollout; defensive)
        first_sub = q.raw_columns[0]
        first_letter = raw_col_to_letter.get(first_sub)
        if not first_letter:
            ws.cell(row=r, column=1, value="(no sub-cols in raw data)")
            return r + 1
        base_range = f"'{raw_sheet_name}'!${first_letter}$2:${first_letter}${n_raw_rows + 1}"
        base_criterion = '"<>"'

    # Pre-compute the row the Base will land on so % column can reference it.
    # After writing `len(q.raw_columns)` option rows starting at r, the next
    # free row IS the Base row.
    base_row = r + len(q.raw_columns)
    base_cell_ref = f"$B${base_row}"

    for sub_col in q.raw_columns:
        letter = raw_col_to_letter.get(sub_col)
        if letter is None:
            continue
        rng = f"'{raw_sheet_name}'!${letter}$2:${letter}${n_raw_rows + 1}"
        label = q.sub_column_labels.get(sub_col, sub_col)
        ws.cell(row=r, column=1, value=str(label)).font = _F_BODY
        formula = (f"=IFERROR(COUNTIFS({rng},1"
                   + (f",{fclause}" if fclause else "")
                   + "),0)")
        ws.cell(row=r, column=2, value=formula).font = _F_BODY
        ws.cell(row=r, column=3, value=f"=IFERROR(B{r}/{base_cell_ref},0)").number_format = "0.0%"
        r += 1
    # Base row
    ws.cell(row=r, column=1, value="Base (any answered)").font = _F_HEAD
    ws.cell(row=r, column=1).fill = _FILL_HEAD
    ws.cell(row=r, column=2, value=(f"=IFERROR(COUNTIFS({base_range},{base_criterion}"
                                     + (f",{fclause}" if fclause else "")
                                     + "),0)")).font = _F_HEAD
    ws.cell(row=r, column=2).fill = _FILL_HEAD
    return r + 1


def _write_grid_rated_block(ws, q, start_row, filter_refs, raw_col_to_letter,
                              raw_sheet_name, n_raw_rows) -> int:
    fclause = _filter_clause(filter_refs, raw_col_to_letter, raw_sheet_name, n_raw_rows)
    ws.cell(row=start_row, column=1, value="Item").font = _F_HEAD
    ws.cell(row=start_row, column=1).fill = _FILL_HEAD
    ws.cell(row=start_row, column=2, value="Valid n").font = _F_HEAD
    ws.cell(row=start_row, column=2).fill = _FILL_HEAD
    ws.cell(row=start_row, column=3, value="Mean").font = _F_HEAD
    ws.cell(row=start_row, column=3).fill = _FILL_HEAD

    r = start_row + 1
    for sub_col in q.raw_columns:
        letter = raw_col_to_letter.get(sub_col)
        if letter is None:
            continue
        rng = f"'{raw_sheet_name}'!${letter}$2:${letter}${n_raw_rows + 1}"
        label = q.sub_column_labels.get(sub_col, sub_col)
        ws.cell(row=r, column=1, value=str(label)).font = _F_BODY
        # Valid n = COUNTIFS(range, ">0", filters)
        ws.cell(row=r, column=2, value=(f"=IFERROR(COUNTIFS({rng},\">0\""
                                          + (f",{fclause}" if fclause else "")
                                          + "),0)"))
        # Mean = SUMIFS / COUNTIFS
        sum_formula = (f"SUMIFS({rng},{rng},\">0\""
                       + (f",{fclause}" if fclause else "")
                       + ")")
        ws.cell(row=r, column=3,
                value=f"=IFERROR({sum_formula}/B{r},0)").number_format = "0.00"
        r += 1
    return r


def _write_numeric_alloc_block(ws, q, start_row, filter_refs, raw_col_to_letter,
                                 raw_sheet_name, n_raw_rows) -> int:
    fclause = _filter_clause(filter_refs, raw_col_to_letter, raw_sheet_name, n_raw_rows)
    ws.cell(row=start_row, column=1, value="Component").font = _F_HEAD
    ws.cell(row=start_row, column=1).fill = _FILL_HEAD
    ws.cell(row=start_row, column=2, value="Valid n").font = _F_HEAD
    ws.cell(row=start_row, column=2).fill = _FILL_HEAD
    ws.cell(row=start_row, column=3, value="Mean %").font = _F_HEAD
    ws.cell(row=start_row, column=3).fill = _FILL_HEAD

    r = start_row + 1
    for sub_col in q.raw_columns:
        letter = raw_col_to_letter.get(sub_col)
        if letter is None:
            continue
        rng = f"'{raw_sheet_name}'!${letter}$2:${letter}${n_raw_rows + 1}"
        label = q.sub_column_labels.get(sub_col, sub_col)
        ws.cell(row=r, column=1, value=str(label)).font = _F_BODY
        ws.cell(row=r, column=2, value=(f"=IFERROR(COUNTIFS({rng},\"<>\""
                                          + (f",{fclause}" if fclause else "")
                                          + "),0)"))
        ws.cell(row=r, column=3, value=(f"=IFERROR(AVERAGEIFS({rng},{rng},\">=0\""
                                          + (f",{fclause}" if fclause else "")
                                          + "),0)")).number_format = "0.0"
        r += 1
    return r


def _write_nps_block(ws, q, start_row, filter_refs, raw_col_to_letter,
                       raw_sheet_name, n_raw_rows) -> int:
    col_letter = raw_col_to_letter.get(q.raw_columns[0])
    if col_letter is None:
        ws.cell(row=start_row, column=1, value=f"(column {q.raw_columns[0]} missing)")
        return start_row + 1
    rng = f"'{raw_sheet_name}'!${col_letter}$2:${col_letter}${n_raw_rows + 1}"
    fclause = _filter_clause(filter_refs, raw_col_to_letter, raw_sheet_name, n_raw_rows)

    ws.cell(row=start_row, column=1, value="Bucket").font = _F_HEAD
    ws.cell(row=start_row, column=1).fill = _FILL_HEAD
    ws.cell(row=start_row, column=2, value="Count").font = _F_HEAD
    ws.cell(row=start_row, column=2).fill = _FILL_HEAD
    ws.cell(row=start_row, column=3, value="%").font = _F_HEAD
    ws.cell(row=start_row, column=3).fill = _FILL_HEAD

    pro_r = start_row + 1
    pas_r = start_row + 2
    det_r = start_row + 3
    nps_r = start_row + 4

    base_formula = (f"COUNTIFS({rng},\">=0\""
                    + (f",{fclause}" if fclause else "")
                    + ")")
    base_cell_ref = f"$B${nps_r + 1}"

    ws.cell(row=pro_r, column=1, value="Promoters (9-10)").font = _F_BODY
    ws.cell(row=pro_r, column=2, value=(f"=IFERROR(COUNTIFS({rng},\">=9\""
                                          + (f",{fclause}" if fclause else "")
                                          + "),0)"))
    ws.cell(row=pro_r, column=3, value=f"=IFERROR(B{pro_r}/{base_cell_ref},0)").number_format = "0.0%"

    ws.cell(row=pas_r, column=1, value="Passives (7-8)").font = _F_BODY
    ws.cell(row=pas_r, column=2, value=(f"=IFERROR(COUNTIFS({rng},\">=7\",{rng},\"<=8\""
                                          + (f",{fclause}" if fclause else "")
                                          + "),0)"))
    ws.cell(row=pas_r, column=3, value=f"=IFERROR(B{pas_r}/{base_cell_ref},0)").number_format = "0.0%"

    ws.cell(row=det_r, column=1, value="Detractors (0-6)").font = _F_BODY
    ws.cell(row=det_r, column=2, value=(f"=IFERROR(COUNTIFS({rng},\"<=6\""
                                          + (f",{fclause}" if fclause else "")
                                          + "),0)"))
    ws.cell(row=det_r, column=3, value=f"=IFERROR(B{det_r}/{base_cell_ref},0)").number_format = "0.0%"

    ws.cell(row=nps_r, column=1, value="NPS").font = _F_HEAD
    ws.cell(row=nps_r, column=1).fill = _FILL_HEAD
    ws.cell(row=nps_r, column=2, value=f"=IFERROR(C{pro_r}*100-C{det_r}*100,0)").number_format = "+0;-0;0"
    ws.cell(row=nps_r, column=2).font = _F_HEAD
    ws.cell(row=nps_r, column=2).fill = _FILL_HEAD

    ws.cell(row=nps_r + 1, column=1, value="Base").font = _F_HEAD
    ws.cell(row=nps_r + 1, column=1).fill = _FILL_HEAD
    ws.cell(row=nps_r + 1, column=2, value=f"=IFERROR({base_formula},0)").font = _F_HEAD
    ws.cell(row=nps_r + 1, column=2).fill = _FILL_HEAD
    return nps_r + 2


def _write_ranking_block(ws, q, start_row, filter_refs, raw_col_to_letter,
                           raw_sheet_name, n_raw_rows) -> int:
    """Ranking layout: options on rows, rank positions (1..K) on columns.

    Each cell = # respondents who assigned that option that specific rank.
    Base = # respondents who ranked at least one option, via the `_q_sum_`
    helper column (same approach as multi-select).
    """
    fclause = _filter_clause(filter_refs, raw_col_to_letter, raw_sheet_name, n_raw_rows)

    # Rank positions come from the question's scale_range (lo..hi).
    # Sensible default if missing: 1..max-observed (here just (1, len(sub_cols))).
    if q.scale_range and q.scale_range[0] is not None:
        rank_lo, rank_hi = q.scale_range
    else:
        rank_lo, rank_hi = 1, len(q.raw_columns) or 1
    rank_values = list(range(int(rank_lo), int(rank_hi) + 1))
    n_ranks = len(rank_values)

    # ── Header row ──
    ws.cell(row=start_row, column=1, value="Item").font = _F_HEAD
    ws.cell(row=start_row, column=1).fill = _FILL_HEAD
    for j, rv in enumerate(rank_values):
        c = 2 + j
        ws.cell(row=start_row, column=c, value=f"Rank {rv}").font = _F_HEAD
        ws.cell(row=start_row, column=c).fill = _FILL_HEAD

    # ── Body rows: one per sub-column (option) ──
    r = start_row + 1
    for sub_col in q.raw_columns:
        letter = raw_col_to_letter.get(sub_col)
        if letter is None:
            continue
        rng = f"'{raw_sheet_name}'!${letter}$2:${letter}${n_raw_rows + 1}"
        label = q.sub_column_labels.get(sub_col, sub_col)
        ws.cell(row=r, column=1, value=str(label)).font = _F_BODY
        for j, rv in enumerate(rank_values):
            c = 2 + j
            formula = (f"=IFERROR(COUNTIFS({rng},{rv}"
                       + (f",{fclause}" if fclause else "")
                       + "),0)")
            ws.cell(row=r, column=c, value=formula).font = _F_BODY
        r += 1

    # ── Base row ──
    helper_header = f"_q_sum_{q.column_id}"
    helper_letter = raw_col_to_letter.get(helper_header)
    ws.cell(row=r, column=1, value="Base (any rank)").font = _F_HEAD
    ws.cell(row=r, column=1).fill = _FILL_HEAD
    if helper_letter:
        base_range = f"'{raw_sheet_name}'!${helper_letter}$2:${helper_letter}${n_raw_rows + 1}"
        base_formula = (f"=IFERROR(COUNTIFS({base_range},\">0\""
                        + (f",{fclause}" if fclause else "")
                        + "),0)")
        ws.cell(row=r, column=2, value=base_formula).font = _F_HEAD
        ws.cell(row=r, column=2).fill = _FILL_HEAD
    return r + 1


def _write_direct_numeric_block(ws, q, start_row, filter_refs, raw_col_to_letter,
                                  raw_sheet_name, n_raw_rows) -> int:
    """Direct-numeric layout: single Mean row. Median/StdDev/Min/Max dropped
    per spec — analysts add those manually if needed."""
    col_letter = raw_col_to_letter.get(q.raw_columns[0])
    if col_letter is None:
        ws.cell(row=start_row, column=1, value=f"(column {q.raw_columns[0]} missing)")
        return start_row + 1
    rng = f"'{raw_sheet_name}'!${col_letter}$2:${col_letter}${n_raw_rows + 1}"
    fclause = _filter_clause(filter_refs, raw_col_to_letter, raw_sheet_name, n_raw_rows)

    ws.cell(row=start_row, column=1, value="Statistic").font = _F_HEAD
    ws.cell(row=start_row, column=1).fill = _FILL_HEAD
    ws.cell(row=start_row, column=2, value="Value").font = _F_HEAD
    ws.cell(row=start_row, column=2).fill = _FILL_HEAD

    ws.cell(row=start_row + 1, column=1, value="Mean").font = _F_BODY
    ws.cell(row=start_row + 1, column=2,
            value=(f"=IFERROR(AVERAGEIFS({rng},{rng},\"<>\""
                   + (f",{fclause}" if fclause else "")
                   + "),0)")).number_format = "0.00"
    return start_row + 2


def _write_cross_cut_sheet(wb: Workbook, cc: CrossCutResult) -> None:
    """Layout (column dim has N categories, "grouped" not "interleaved"):

        Row 1     : merged title
        Row 3-4   : Row / Col dimension labels
        Row 6     : "# of respondents" merged across N count cols,
                    "% of respondents" merged across N pct cols
        Row 7     : option label 1..N (under #), option label 1..N (under %)
        Row 8+    : row label | c_1 c_2 .. c_N | p_1 p_2 .. p_N
        Last+1    : "Total"   | sum_1 .. sum_N | 100% .. 100%

    The % base is the COLUMN TOTAL (sum down each count column). Each cell
    percent equals cell_count / column_total.
    """
    used = set(wb.sheetnames)
    name = f"X · {cc.row_column_id[:10]} × {cc.col_column_id[:10]}"
    sn = _safe_sheet(name, used)
    ws = wb.create_sheet(sn)
    ws.column_dimensions["A"].width = 40
    n_cols = len(cc.col_labels)
    total_cols = 1 + 2 * n_cols  # col A label + N count cols + N % cols
    for c in range(2, total_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 16

    # Column index helpers — counts live at columns [2 .. 1+n_cols],
    # percents at columns [2+n_cols .. 1+2*n_cols].
    count_col_start = 2
    count_col_end = 1 + n_cols
    pct_col_start = count_col_end + 1
    pct_col_end = pct_col_start + n_cols - 1

    # ── Title ──
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=total_cols)
    ws.cell(row=1, column=1,
            value=f"{cc.row_question_text[:60]} × {cc.col_question_text[:60]}").font = _F_THEME_TITLE

    ws.cell(row=3, column=1, value=f"Row: {cc.row_column_id}").font = _F_SECTION
    ws.cell(row=4, column=1, value=f"Col: {cc.col_column_id}").font = _F_SECTION

    # ── Row 6 — top header: "# of respondents" / "% of respondents" group titles ──
    base_r = 6
    ws.cell(row=base_r, column=1, value="").fill = _FILL_HEAD
    # # group
    num_cell = ws.cell(row=base_r, column=count_col_start, value="# of respondents")
    num_cell.font = _F_HEAD
    num_cell.fill = _FILL_HEAD
    num_cell.alignment = ALIGN_CENTER
    if n_cols > 1:
        ws.merge_cells(start_row=base_r, end_row=base_r,
                       start_column=count_col_start, end_column=count_col_end)
    for c in range(count_col_start + 1, count_col_end + 1):
        ws.cell(row=base_r, column=c).fill = _FILL_HEAD
    # % group
    pct_cell = ws.cell(row=base_r, column=pct_col_start, value="% of respondents")
    pct_cell.font = _F_HEAD
    pct_cell.fill = _FILL_HEAD
    pct_cell.alignment = ALIGN_CENTER
    if n_cols > 1:
        ws.merge_cells(start_row=base_r, end_row=base_r,
                       start_column=pct_col_start, end_column=pct_col_end)
    for c in range(pct_col_start + 1, pct_col_end + 1):
        ws.cell(row=base_r, column=c).fill = _FILL_HEAD

    # ── Row 7 — sub-header: option labels under both # and % groups ──
    hdr_r = base_r + 1
    ws.cell(row=hdr_r, column=1, value="").fill = _FILL_HEAD
    for j, lbl in enumerate(cc.col_labels):
        for col_offset in (count_col_start, pct_col_start):
            c_idx = col_offset + j
            cell = ws.cell(row=hdr_r, column=c_idx, value=str(lbl)[:60])
            cell.font = _F_HEAD
            cell.fill = _FILL_HEAD
            cell.alignment = ALIGN_CENTER

    # ── Body rows ──
    # Pre-compute column totals
    col_totals = [0.0] * n_cols
    for row_vals in cc.counts:
        for j, v in enumerate(row_vals):
            try:
                col_totals[j] += float(v)
            except (TypeError, ValueError):
                pass

    body_start = base_r + 2
    for i, rlbl in enumerate(cc.row_labels):
        r = body_start + i
        lbl_cell = ws.cell(row=r, column=1, value=str(rlbl)[:80])
        lbl_cell.font = _F_HEAD
        lbl_cell.fill = _FILL_HEAD
        for j, val in enumerate(cc.counts[i]):
            # count
            ws.cell(row=r, column=count_col_start + j, value=val).font = _F_BODY
            # %
            denom = col_totals[j]
            pct = (float(val) / denom) if denom else 0.0
            pct_cell = ws.cell(row=r, column=pct_col_start + j, value=pct)
            pct_cell.number_format = "0.0%"
            pct_cell.font = _F_PCT

    # ── Total row ──
    total_r = body_start + len(cc.row_labels)
    tot_cell = ws.cell(row=total_r, column=1, value="Total")
    tot_cell.font = _F_HEAD
    tot_cell.fill = _FILL_HEAD
    for j in range(n_cols):
        # count total
        ws.cell(row=total_r, column=count_col_start + j, value=col_totals[j]).font = _F_HEAD
        ws.cell(row=total_r, column=count_col_start + j).fill = _FILL_HEAD
        # % total = 100% by definition
        pcell = ws.cell(row=total_r, column=pct_col_start + j,
                        value=1.0 if col_totals[j] else 0.0)
        pcell.number_format = "0.0%"
        pcell.font = _F_HEAD
        pcell.fill = _FILL_HEAD


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────


def _excel_str(value: Any) -> str:
    """Quote a value for use inside an Excel formula."""
    if isinstance(value, str):
        # Escape any inner quotes
        safe = value.replace('"', '""')
        return f'"{safe}"'
    return str(value)


def _reorder_tabs(
    wb: Workbook,
    themes: list[ThemeGroup],
    cross_cuts: list[CrossCutResult],
    raw_sheet_name: str,
) -> None:
    """Make the tab order: Output>>, themes..., Mapping>>, Datamap, Validation, Data>>, Raw data, cross cuts..."""
    desired: list[str] = []
    desired.append("Output>>")
    desired.extend(t.name[:31] for t in themes)
    desired.append("Mapping>>")
    desired.append("Datamap")
    desired.append("Validation")
    desired.append("Data>>")
    desired.append(raw_sheet_name)
    for cc in cross_cuts:
        candidate = f"X · {cc.row_column_id[:10]} × {cc.col_column_id[:10]}"
        # Find the actual created sheet name (theme grouper / safe_sheet may have renamed)
        for sn in wb.sheetnames:
            if sn.startswith("X ") and cc.row_column_id[:10] in sn and cc.col_column_id[:10] in sn:
                if sn not in desired:
                    desired.append(sn)
                break

    # Build new sheet order, keeping only sheets that exist
    new_order: list[str] = []
    for sn in desired:
        if sn in wb.sheetnames and sn not in new_order:
            new_order.append(sn)
    # Append any sheets we forgot
    for sn in wb.sheetnames:
        if sn not in new_order:
            new_order.append(sn)
    # Apply via _sheets reordering (openpyxl supports this)
    wb._sheets = [wb[sn] for sn in new_order]
