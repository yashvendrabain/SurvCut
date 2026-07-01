"""End-to-end smoke test — runs the full Phase 2 pipeline on the sample file.

Stages:
  1. Load combined .xlsx
  2. Parse + classify + validate
  3. Compute all single cuts
  4. Compute a dynamic cross cut
  5. Suggest themes + filters
  6. Build the output workbook
  7. Reopen it and report tab structure
"""
from __future__ import annotations

import os
import shutil
import sys
import tempfile
from pathlib import Path

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))

from openpyxl import load_workbook

from core.classifier import classify, summarise
from core.cross_cut import compute_cross_cut
from core.datamap_parser import parse_datamap_from_rows
from core.exporter import ExportInputs, export
from core.io_layer import load_combined
from core.models import FilterSlot, QuestionType
from core.single_cut import compute_all_single_cuts
from core.theme_grouper import suggest_themes
from core.validator import format_report, validate


ORIGINAL = (
    HERE.parent
    / "Survey cutter automation"
    / "Survey insight engine runnable"
    / "docs"
    / "sample_datamap_and_rawdata.xlsx"
)


def _readable_path() -> Path:
    """OneDrive can transiently lock the original — read via a %TEMP% copy."""
    tmp = Path(tempfile.gettempdir()) / "sample_dm_copy.xlsx"
    try:
        shutil.copy(ORIGINAL, tmp)
        return tmp
    except (PermissionError, OSError):
        if tmp.exists():
            return tmp
        return ORIGINAL


def section(t: str) -> None:
    print(f"\n{'=' * 70}\n{t}\n{'=' * 70}")


def main() -> None:
    sample = _readable_path()
    print(f"Sample: {sample}")

    section("STAGE 1 — Load combined .xlsx")
    loaded = load_combined(sample)
    print(f"  raw: {loaded.raw_df.shape}  datamap rows: {len(loaded.datamap_rows)}")
    print(f"  {loaded.raw_source_note}  |  {loaded.datamap_source_note}")

    section("STAGE 2 — Parse + classify + validate")
    blocks = parse_datamap_from_rows(loaded.datamap_rows)
    schema = classify(blocks, loaded.raw_df)
    report = validate(blocks, loaded.raw_df)
    print(f"  Parsed {len(blocks)} blocks → schema has {len(schema.questions)} questions.")
    print("  " + format_report(report).replace("\n", "\n  "))

    section("STAGE 3 — Compute all single cuts")
    cuts = compute_all_single_cuts(schema, loaded.raw_df)
    print(f"  Computed {len(cuts)} single cuts.")
    for c in cuts[:4]:
        head = c.headline_metric or "(no headline)"
        print(f"    {c.column_id:<28} {c.question_type.value:<22} "
              f"n={c.valid_n:>4}  rows={len(c.rows):>2}  | {head}")

    section("STAGE 4 — Dynamic cross-cut: Q7_Sector × Q15_Recommend (NPS)")
    q_sector = schema.by_column_id("Q7_Sector")
    q_nps = schema.by_column_id("Q15_Recommend")
    xc = compute_cross_cut(q_sector, q_nps, loaded.raw_df)
    print(f"  Cross matrix: {len(xc.row_labels)} rows × {len(xc.col_labels)} cols")
    print(f"  Row labels: {list(xc.row_labels)}")
    print(f"  Col labels: {list(xc.col_labels)}")
    print(f"  Counts:")
    for rlbl, row_vals in zip(xc.row_labels, xc.counts):
        print(f"    {rlbl:<24} {row_vals}")
    print(f"  Grand total: {xc.grand_total}")

    section("STAGE 5 — Suggest themes + filters")
    themes = suggest_themes(schema)
    for t in themes:
        print(f"  {t.name:<35}  {len(t.question_column_ids)} questions")
    # Pick the first three demographic single-select as filter slots
    demo_filters = [
        FilterSlot(name=q.column_id, column_id=q.column_id, default_value="All")
        for q in schema.questions
        if q.analysis_eligible and q.is_demographic
        and q.question_type == QuestionType.SINGLE_SELECT
    ][:3]
    print(f"  Filters: {[f.column_id for f in demo_filters]}")

    section("STAGE 6 — Build output workbook")
    out_path = os.path.join(tempfile.gettempdir(), "cutter_v2_smoke.xlsx")
    inputs = ExportInputs(
        schema=schema, raw_df=loaded.raw_df, datamap_rows=loaded.datamap_rows,
        themes=themes, filters=demo_filters, cross_cuts=(xc,),
    )
    path = export(inputs, out_path)
    size = Path(path).stat().st_size
    print(f"  Wrote: {path}  ({size:,} bytes)")

    section("STAGE 7 — Reopen and inspect")
    wb = load_workbook(path, data_only=False)
    print("  Sheet order:")
    for i, sn in enumerate(wb.sheetnames):
        ws = wb[sn]
        print(f"    {i+1:>2}. {sn:<32}  {ws.max_row}r × {ws.max_column}c")
    # Sample a formula from each theme sheet
    print("\n  Sample formulas (first 3 from each theme):")
    for sn in wb.sheetnames:
        if sn in ("Output>>", "Mapping>>", "Data>>", "Raw data", "Datamap", "Validation"):
            continue
        if sn.startswith("X "):
            continue
        ws = wb[sn]
        fcount = 0
        for row in ws.iter_rows(min_row=15, max_row=min(60, ws.max_row), max_col=4):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    val = cell.value.replace("\n", " ")
                    print(f"    [{sn}!{cell.coordinate}] {val[:140]}")
                    fcount += 1
                    if fcount >= 3:
                        break
            if fcount >= 3:
                break

    section("DONE")
    print("Open the file in Excel to inspect:")
    print(f"  {out_path}")


if __name__ == "__main__":
    main()
